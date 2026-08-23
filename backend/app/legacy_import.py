from __future__ import annotations

import copy
import hashlib
import json
import math
import mimetypes
import os
import re
import secrets
import shutil
import urllib.parse
import urllib.request
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timezone
from html import unescape
from html.parser import HTMLParser
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.engine import make_url
from sqlalchemy.orm import Session
from sqlalchemy.orm.attributes import flag_modified

from app.media_service import _matches_declared_content, media_access_reference, normalize_content_type
from app.models.audit import LegacyImportRecord
from app.models.board import Board
from app.models.comment import Comment
from app.models.faq import FAQ, FAQAttachment
from app.models.media import MediaAsset, PostAttachment
from app.models.post import Post
from app.models.post_extension import PostSuggestion
from app.models.user import User
from app.security import hash_password, utc_now
from app.study_activity_cleanup import (
    LEGACY_STUDY_ACTIVITY_SOURCE_WRITE_IDS,
    LEGACY_STUDY_ACTIVITY_TITLES,
    curated_study_activity_title,
)


ARTICLE_HEADERS = [
    "boardName", "writeId", "parentWriteId", "boardId", "writeType", "title", "content",
    "writeUser", "cohort", "writeUserStr", "writeUserDbId", "writeUserImage",
    "writeUserImageUrl", "date", "updateDate", "likeCount", "readCount", "viewCount",
    "commentCount", "pushCount", "appId", "appName", "stat", "sequence", "isLike",
    "hashTagListStr", "isFirstImgThumbnail", "reserveDateTime", "ogStat", "startDatetime",
    "endDatetime", "calBackColor", "calTextColor", "boardArticleMetaList", "reportList",
    "notices_미포함",
]

SHEET_TO_BOARD_SLUG = {
    "학사공지": "academic-notices",
    "행사공지": "event-notices",
    "기타공지": "all-notices",
    "강의후기": "lecture-reviews",
    "시험족보": "exam-archive",
    "종합시험": "comprehensive-exam",
    "졸업논문": "graduation-thesis",
    "동아리 안내": "club-promo",
    "동아리 활동 인증": "club-activity",
    "스터디 모집": "study-recruit",
    "스터디 활동 인증": "study-activity",
    "사진첩": "event-album",
    "건의사항": "suggestions",
    "건의사항 피드백": "gsa-feedback",
}

STAGING_BOARD_TO_SLUG = {
    "전공 커뮤니티": "community-major",
    "세미나 공유": "community-seminar",
    "채용정보": "community-job",
    "채용 정보": "community-job",
    "자료 공유": "exam-archive",
    "논문 자료": "community-paper",
}

SPECIAL_SHEETS = {
    "역대 원우회": "past_council",
    "기수별 기장단 소개": "cohort_leader",
    "원우회 상조회": "mutual_aid_archive",
    "자주 묻는 질문": "faq",
}

COMMENT_DISABLED_SLUGS = {
    "lecture-reviews",
    "academic-notices",
    "event-notices",
    "all-notices",
    "webinar-notices",
    "club-activity",
    "study-activity",
    "networking-activity",
    "suggestions",
    "gsa-feedback",
    "mutual-aid",
    "gsa-cohort-leaders",
    "gsa-past-councils",
}

NOTICE_CATEGORY = {
    "학사공지": "academic",
    "행사공지": "event",
    "기타공지": "other",
}

EMAIL_RE = re.compile(r"(?<![\w.+-])[\w.+-]+@[\w.-]+\.[A-Za-z]{2,}(?![\w.-])")
PHONE_RE = re.compile(r"(?<!\d)(?:01[016789])[-.\s]?\d{3,4}[-.\s]?\d{4}(?!\d)")
STUDENT_ID_RE = re.compile(r"(?<!\d)(?:19|20)\d{6,8}(?!\d)")
LEGACY_STUDENT_ID_RE = re.compile(r"(?<![A-Za-z0-9])A\d{5,8}(?![A-Za-z0-9])", re.I)
PLACEHOLDER_TITLE_RE = re.compile(r"^\s*\[?\s*강의\s*명\s*-\s*\d{4}년도\s*\d\s*학기\s*\]?\s*", re.I)

ALLOWED_MIME_TYPES = {
    "image/jpeg", "image/png", "image/gif", "image/webp", "image/heic", "image/heif",
    "application/pdf", "application/msword", "application/vnd.ms-excel",
    "application/vnd.ms-powerpoint",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.openxmlformats-officedocument.presentationml.presentation",
    "application/x-hwp", "application/haansofthwp", "application/vnd.hancom.hwp",
    "application/zip", "text/plain", "application/x-ipynb+json", "application/json",
}

MIME_EXTENSIONS = {
    "image/jpeg": ".jpg",
    "image/png": ".png",
    "image/gif": ".gif",
    "image/webp": ".webp",
    "image/heic": ".heic",
    "image/heif": ".heif",
    "application/pdf": ".pdf",
    "application/msword": ".doc",
    "application/vnd.ms-excel": ".xls",
    "application/vnd.ms-powerpoint": ".ppt",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document": ".docx",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": ".xlsx",
    "application/vnd.openxmlformats-officedocument.presentationml.presentation": ".pptx",
    "application/x-hwp": ".hwp",
    "application/haansofthwp": ".hwp",
    "application/vnd.hancom.hwp": ".hwp",
    "application/zip": ".zip",
    "text/plain": ".txt",
    "application/x-ipynb+json": ".ipynb",
    "application/json": ".ipynb",
}

# The cleaned workbook contains one article ID in two sheets.  Its board name
# and attachments both identify the photo album row as the canonical record.
# Keeping this decision explicit prevents workbook sheet order from silently
# routing the post and its attachments to the wrong board.
CANONICAL_DUPLICATE_ARTICLE_SHEETS = {
    "6764961": "사진첩",
}

# Video remains outside the member attachment allowlist. Archive, text, and
# notebook files are validated and restored through the normal media flow.
ARCHIVED_LEGACY_ATTACHMENT_EXTENSIONS = {".mp4"}


@dataclass(frozen=True)
class SourceRow:
    source_file: str
    sheet: str
    row_number: int
    data: dict[str, Any]

    @property
    def source_id(self) -> str:
        return value_as_text(self.data.get("writeId"))

    @property
    def source_hash(self) -> str:
        return row_hash(self.data)


class _HTMLTextExtractor(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.parts: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag in {"br", "p", "div", "li", "tr"}:
            self.parts.append("\n")
        if tag == "a":
            href = dict(attrs).get("href")
            if href:
                self.parts.append(f" {href} ")

    def handle_endtag(self, tag: str) -> None:
        if tag in {"p", "div", "li", "tr"}:
            self.parts.append("\n")

    def handle_data(self, data: str) -> None:
        self.parts.append(data)


def value_as_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def value_as_int(value: Any, default: int = 0) -> int:
    text = value_as_text(value).replace(",", "")
    if not text:
        return default
    try:
        return int(float(text))
    except ValueError:
        return default


def parse_datetime(value: Any) -> datetime | None:
    if value is None or value_as_text(value) == "":
        return None
    if isinstance(value, datetime):
        parsed = value
    else:
        text = value_as_text(value).replace("Z", "+00:00")
        parsed = None
        for date_format in (
            "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y.%m.%d %H:%M:%S",
            "%Y.%m.%d %H:%M", "%Y.%m.%d %H:%M:%S.%f",
        ):
            try:
                parsed = datetime.strptime(text, date_format)
                break
            except ValueError:
                continue
        if parsed is None:
            try:
                parsed = datetime.fromisoformat(text)
            except ValueError:
                return None
    if parsed.tzinfo is not None:
        parsed = parsed.astimezone(timezone.utc).replace(tzinfo=None)
    return parsed


def normalize_content(value: Any) -> str:
    text = value_as_text(value)
    if "<" not in text or ">" not in text:
        return re.sub(r"\n{3,}", "\n\n", text).strip()
    parser = _HTMLTextExtractor()
    parser.feed(text)
    extracted = unescape("".join(parser.parts))
    extracted = re.sub(r"[ \t]+\n", "\n", extracted)
    return re.sub(r"\n{3,}", "\n\n", extracted).strip()


def normalize_cohort(value: Any) -> str | None:
    match = re.search(r"(?<!\d)(\d{2})\s*기?", value_as_text(value))
    return f"{match.group(1)}기" if match else None


def normalize_author(write_user: Any, cohort: Any = None) -> tuple[str, str | None]:
    raw = value_as_text(write_user).strip("[] ")
    normalized_cohort = normalize_cohort(cohort) or normalize_cohort(raw)
    name = raw
    name = re.sub(r"^서강대(?:학교)?\s*", "", name)
    if normalized_cohort:
        cohort_number = normalized_cohort[:-1]
        name = re.sub(rf"(?:^|[\s_\-]){re.escape(cohort_number)}\s*기(?=$|[\s_\-])", " ", name)
        name = re.sub(rf"^(?:{re.escape(cohort_number)}\s*기[\s_\-]*)+", "", name)
    name = re.sub(r"^[\s_\-]+|[\s_\-]+$", "", name)
    name = re.sub(r"\s{2,}", " ", name)
    return (name or "과거 작성자")[:50], normalized_cohort


def redact_text(value: Any) -> tuple[str, list[str]]:
    text = value_as_text(value)
    findings: list[str] = []
    if EMAIL_RE.search(text):
        findings.append("email")
        text = EMAIL_RE.sub("[이메일 비공개]", text)
    if PHONE_RE.search(text):
        findings.append("phone")
        text = PHONE_RE.sub("[연락처 비공개]", text)
    if STUDENT_ID_RE.search(text) or LEGACY_STUDENT_ID_RE.search(text):
        findings.append("student_id")
        text = STUDENT_ID_RE.sub("[학번 비공개]", text)
        text = LEGACY_STUDENT_ID_RE.sub("[학번 비공개]", text)
    return text, findings


def row_hash(data: dict[str, Any]) -> str:
    rendered = json.dumps(data, ensure_ascii=False, sort_keys=True, default=str, separators=(",", ":"))
    return hashlib.sha256(rendered.encode("utf-8")).hexdigest()


def shortened_title(title: str, maximum: int = 100) -> str:
    title = PLACEHOLDER_TITLE_RE.sub("", title).strip() or "제목 없음"
    if len(title) <= maximum:
        return title
    return title[: maximum - 1].rstrip() + "…"


def _nonempty_rows(sheet) -> Iterable[tuple[int, tuple[Any, ...]]]:
    for row_number, values in enumerate(sheet.iter_rows(values_only=True), start=1):
        if any(value not in (None, "") for value in values):
            yield row_number, values


def load_article_workbook(path: Path) -> tuple[list[SourceRow], list[SourceRow], list[SourceRow]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    articles: list[SourceRow] = []
    attachments: list[SourceRow] = []
    duplicates: list[SourceRow] = []
    seen_ids: set[str] = set()
    article_indexes: dict[str, int] = {}
    for sheet in workbook.worksheets:
        rows = list(_nonempty_rows(sheet))
        if not rows:
            continue
        first_values = [value_as_text(value) for value in rows[0][1]]
        has_header = "writeId" in first_values
        if sheet.title == "첨부파일":
            headers = first_values if has_header else [
                "writeId", "boardId", "attach_type", "sequence", "subject", "fileStorageId",
                "contentType", "attach_id", "link_url", "regiDatetime", "fileSize",
            ]
            for row_number, values in rows[1:] if has_header else rows:
                data = {headers[index]: value for index, value in enumerate(values) if index < len(headers)}
                attachments.append(SourceRow(path.name, sheet.title, row_number, data))
            continue
        headers = first_values if has_header else ARTICLE_HEADERS
        for row_number, values in rows[1:] if has_header else rows:
            data = {headers[index]: value for index, value in enumerate(values) if index < len(headers)}
            row = SourceRow(path.name, sheet.title, row_number, data)
            if not row.source_id:
                continue
            if row.source_id in seen_ids:
                canonical_sheet = CANONICAL_DUPLICATE_ARTICLE_SHEETS.get(row.source_id)
                existing_index = article_indexes[row.source_id]
                existing_row = articles[existing_index]
                if canonical_sheet and row.sheet == canonical_sheet and existing_row.sheet != canonical_sheet:
                    duplicates.append(existing_row)
                    articles[existing_index] = row
                else:
                    duplicates.append(row)
                continue
            seen_ids.add(row.source_id)
            article_indexes[row.source_id] = len(articles)
            articles.append(row)
    return articles, attachments, duplicates


def load_comment_workbook(path: Path) -> list[SourceRow]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(_nonempty_rows(sheet))
    if not rows:
        return []
    headers = [value_as_text(value) for value in rows[0][1]]
    result = []
    for row_number, values in rows[1:]:
        data = {headers[index]: value for index, value in enumerate(values) if index < len(headers)}
        data["writeId"] = data.get("comment_id")
        result.append(SourceRow(path.name, sheet.title, row_number, data))
    return result


def load_reference_attachment_urls(path: Path) -> dict[str, str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if "첨부파일" not in workbook.sheetnames:
        return {}
    sheet = workbook["첨부파일"]
    rows = list(_nonempty_rows(sheet))
    headers = [value_as_text(value) for value in rows[0][1]]
    result: dict[str, str] = {}
    for _, values in rows[1:]:
        data = {headers[index]: value for index, value in enumerate(values) if index < len(headers)}
        storage_id = value_as_text(data.get("fileStorageId"))
        url = value_as_text(data.get("url"))
        if storage_id and url.startswith(("http://", "https://")):
            result.setdefault(storage_id, url)
    return result


def target_slug(row: SourceRow) -> str | None:
    if row.sheet == "자유게시판":
        return STAGING_BOARD_TO_SLUG.get(value_as_text(row.data.get("boardName")))
    return SHEET_TO_BOARD_SLUG.get(row.sheet)


def validate_database_target(database_url: str) -> str:
    database_name = make_url(database_url).database or ""
    if database_name == "sogang_app":
        raise RuntimeError("Refusing to import into the protected source database 'sogang_app'.")
    if "migration_review" not in database_name and not database_name.startswith("test_"):
        raise RuntimeError(
            "Import target must be an isolated migration_review or test database; "
            f"received {database_name!r}."
        )
    return database_name


def _existing_posts_by_legacy_id(db: Session) -> dict[str, Post]:
    result: dict[str, Post] = {}
    for post in db.scalars(select(Post)).all():
        metadata = post.metadata_json or {}
        source_id = metadata.get("legacy_write_id") or metadata.get("legacy_article_id")
        if source_id is not None:
            result[str(source_id)] = post
    return result


def _legacy_user_cache(db: Session) -> dict[tuple[str, str], list[User]]:
    cache: dict[tuple[str, str], list[User]] = defaultdict(list)
    for user in db.scalars(select(User).where(User.username.like("legacy%"))).all():
        name, cohort = normalize_author(user.nickname, user.cohort)
        cache[(name, cohort or "")].append(user)
    return cache


def ensure_legacy_user(
    db: Session,
    write_user: Any,
    cohort: Any,
    cache: dict[tuple[str, str], list[User]],
    preferred: User | None = None,
) -> User:
    name, normalized_cohort = normalize_author(write_user, cohort)
    if preferred is not None and preferred.username.startswith("legacy") and not preferred.is_active:
        preferred.nickname = name
        preferred.cohort = normalized_cohort
        cache.setdefault((name, normalized_cohort or ""), []).append(preferred)
        return preferred
    candidates = cache.get((name, normalized_cohort or ""), [])
    if len(candidates) == 1:
        return candidates[0]
    digest = hashlib.sha256(f"{normalized_cohort or ''}|{name}".encode("utf-8")).hexdigest()[:24]
    username = f"legacyv2_{digest}"
    existing = db.scalar(select(User).where(User.username == username))
    if existing is None:
        existing = User(
            username=username,
            email=f"legacyv2-{digest}@invalid.local",
            password_hash=hash_password(secrets.token_urlsafe(48)),
            nickname=name,
            cohort=normalized_cohort,
            role="user",
            is_active=False,
        )
        db.add(existing)
        db.flush()
    cache.setdefault((name, normalized_cohort or ""), []).append(existing)
    return existing


def upsert_ledger(
    db: Session,
    row: SourceRow,
    *,
    entity_type: str,
    action: str,
    status: str,
    source_parent_id: str | None = None,
    target_table: str | None = None,
    target_id: int | None = None,
    reason: str | None = None,
    details: dict[str, Any] | None = None,
) -> LegacyImportRecord:
    record = db.scalar(
        select(LegacyImportRecord).where(
            LegacyImportRecord.source_file == row.source_file,
            LegacyImportRecord.source_sheet == row.sheet,
            LegacyImportRecord.entity_type == entity_type,
            LegacyImportRecord.source_id == row.source_id,
        )
    )
    payload = {
        "source_row": row.row_number,
        "source_parent_id": source_parent_id,
        "source_hash": row.source_hash,
        "action": action,
        "status": status,
        "target_table": target_table,
        "target_id": target_id,
        "reason": reason,
        "redacted_details": details,
    }
    if record is None:
        record = LegacyImportRecord(
            source_file=row.source_file,
            source_sheet=row.sheet,
            entity_type=entity_type,
            source_id=row.source_id,
            **payload,
        )
        db.add(record)
        db.flush()
        return record
    if action == "unchanged":
        payload["action"] = record.action
    changed = False
    for key, value in payload.items():
        if getattr(record, key) != value:
            setattr(record, key, value)
            changed = True
    if changed:
        record.updated_at = utc_now()
    return record


def _post_payload(
    row: SourceRow,
    board: Board,
    author: User,
    *,
    activity_source_post_id: int | None = None,
) -> dict[str, Any]:
    source_title = value_as_text(row.data.get("title")) or "제목 없음"
    original_title = (
        curated_study_activity_title(row.source_id, source_title)
        if board.slug == "study-activity"
        else source_title
    )
    normalized_body = normalize_content(row.data.get("content")) or original_title
    redacted_source_title, title_pii = redact_text(source_title)
    uses_curated_study_title = (
        board.slug == "study-activity"
        and row.source_id in LEGACY_STUDY_ACTIVITY_TITLES
    )
    redacted_title = original_title if uses_curated_study_title else redacted_source_title
    redacted_body, body_pii = redact_text(normalized_body)
    title = shortened_title(redacted_title)
    created_at = parse_datetime(row.data.get("date"))
    if created_at is None:
        raise ValueError("missing_created_at")
    updated_at = parse_datetime(row.data.get("updateDate")) or created_at
    metadata = {
        "legacy_source": "swing2app_ver2",
        "legacy_sheet": row.sheet,
        "legacy_write_id": row.source_id,
        "legacy_board_name": value_as_text(row.data.get("boardName")),
        "legacy_author": author.nickname,
        "legacy_author_cohort": author.cohort,
        "legacy_original_title": redacted_source_title,
        "legacy_source_hash": row.source_hash,
        "legacy_updated_at": updated_at.isoformat(),
        "legacy_comment_count": value_as_int(row.data.get("commentCount")),
        "legacy_pii_redacted": sorted(set(title_pii + body_pii)),
    }
    if board.board_type == "activity_certification":
        metadata.update(_activity_certification_metadata(normalized_body, created_at, title))
        if activity_source_post_id is not None:
            metadata["activity_source_post_id"] = str(activity_source_post_id)
    if board.slug == "study-recruit":
        metadata["recruitment_status"] = "closed"
    if board.board_type == "notice":
        category = NOTICE_CATEGORY.get(row.sheet)
    elif row.sheet == "자유게시판":
        # The workbook sheet is a staging bucket, not an end-user category.
        category = board.name
    else:
        category = row.sheet
    return {
        "board_id": board.id,
        "author_id": author.id,
        "author_nickname_snapshot": author.nickname,
        "author_cohort_snapshot": author.cohort,
        "title": title,
        "content": redacted_body,
        "is_anonymous": board.slug == "lecture-reviews",
        "is_notice": board.board_type == "notice",
        "status": "published",
        "category": category,
        "metadata_json": metadata,
        "view_count": value_as_int(row.data.get("viewCount")),
        "like_count": value_as_int(row.data.get("likeCount")),
        "created_at": created_at,
        "updated_at": updated_at,
        "deleted_at": None,
    }


def _activity_certification_metadata(content: str, created_at: datetime, title: str) -> dict[str, str]:
    lines = [line.strip() for line in content.splitlines()]
    participant_labels = {"[이름 / 기수]", "[이름/기수]", "[참가자]", "[참여자]", "[참여 인원]"}
    date_labels = {"[활동 날짜]", "[활동날짜]", "[활동 일자]", "[활동일자]", "[첫 활동일]"}
    participants: list[str] = []
    activity_date = created_at.date()

    def add_participant(name: str, cohort_number: str = "") -> None:
        name = name.strip()
        if not re.fullmatch(r"[가-힣]{2,5}", name):
            return
        cohort = f"{cohort_number}기" if cohort_number else ""
        label = f"{cohort} {name}".strip()
        if label not in participants:
            participants.append(label)

    def parsed_date(candidate: str) -> date | None:
        match = re.search(
            r"(?:(?P<year>\d{2,4})\s*(?:년|[./-])\s*)?"
            r"(?P<month>\d{1,2})\s*(?:월|[./-])\s*(?P<day>\d{1,2})",
            candidate,
        )
        if not match:
            return None
        try:
            year = int(match.group("year") or created_at.year)
            if year < 100:
                year += 2000
            return date(year, int(match.group("month")), int(match.group("day")))
        except ValueError:
            return None

    for index, line in enumerate(lines):
        if line in participant_labels or line.startswith("[스터디원 이름"):
            for candidate in lines[index + 1:]:
                if candidate.startswith("[") and candidate.endswith("]"):
                    if any(
                        marker in candidate
                        for marker in ("활동 날짜", "첫 활동일", "이미지", "스터디 날짜", "스터디 내용")
                    ):
                        break
                    continue
                if (
                    not candidate
                    or re.search(r"총\s*\d+\s*명", candidate)
                    or re.search(r"\d{1,2}\s*[./월]\s*\d{1,2}", candidate)
                    or "모임 참석" in candidate
                ):
                    continue
                legacy_student_match = re.match(
                    r"^(?:\d+\.\s*)?A(?P<cohort>\d{2})\d{3,}\s+(?P<name>[가-힣]{2,5})",
                    candidate,
                    re.I,
                )
                grouped_cohort_match = re.match(
                    r"^(?:\d+\.\s*)?(?P<cohort>\d{2})기?\s*[-:]\s*(?P<names>.+)$",
                    candidate,
                )
                participant_match = re.match(
                    r"^(?:\d+\.\s*)?(?:(?P<cohort_first>\d{2})(?:기)?[\s_/]+(?P<name_after>[가-힣]{2,5})|"
                    r"(?P<name_first>[가-힣]{2,5})[\s_/]+(?P<cohort_after>\d{2})(?:기)?)",
                    candidate,
                )
                if legacy_student_match:
                    add_participant(legacy_student_match.group("name"), legacy_student_match.group("cohort"))
                    continue
                if grouped_cohort_match:
                    cohort_number = grouped_cohort_match.group("cohort")
                    for name_part in re.split(r"[,，]", grouped_cohort_match.group("names")):
                        name_match = re.search(r"[가-힣]{2,5}", name_part)
                        if name_match:
                            add_participant(name_match.group(), cohort_number)
                    continue
                elif participant_match:
                    cohort_number = participant_match.group("cohort_first") or participant_match.group("cohort_after") or ""
                    name = participant_match.group("name_after") or participant_match.group("name_first") or ""
                else:
                    name, normalized_cohort = normalize_author(candidate)
                    cohort_number = (normalized_cohort or "").removesuffix("기")
                add_participant(name, cohort_number)
            break

    # Older study forms contain only a paid-member student ID next to each
    # participant. The ID is used solely to recover the cohort and is never
    # retained in the imported metadata.
    legacy_identity_pattern = re.compile(
        r"(?:(?P<name_first>[가-힣]{2,5})\s*\(\s*A(?P<cohort_after>\d{2})\d{3,}\s*\)|"
        r"A(?P<cohort_first>\d{2})\d{3,}\s+(?P<name_after>[가-힣]{2,5}))",
        re.I,
    )
    for match in legacy_identity_pattern.finditer(content):
        add_participant(
            match.group("name_first") or match.group("name_after") or "",
            match.group("cohort_after") or match.group("cohort_first") or "",
        )

    for index, line in enumerate(lines):
        if line not in date_labels and not line.startswith("[스터디 날짜"):
            continue
        candidate = next(
            (
                value
                for value in lines[index + 1:index + 5]
                if value and not (value.startswith("[") and value.endswith("]"))
            ),
            "",
        )
        extracted_date = parsed_date(candidate)
        if extracted_date:
            activity_date = extracted_date
        break

    if activity_date == created_at.date():
        meeting_date = re.search(r"모임일자\s*[:：]\s*([^\r\n]+)", content)
        extracted_date = parsed_date(meeting_date.group(1) if meeting_date else content)
        if extracted_date:
            activity_date = extracted_date

    metadata = {
        "activity_date": activity_date.isoformat(),
        "legacy_activity_name": title,
    }
    if participants:
        metadata["participants"] = ", ".join(participants)
    return metadata


def _assign_if_changed(target: Any, payload: dict[str, Any]) -> bool:
    changed = False
    for key, value in payload.items():
        if getattr(target, key) != value:
            setattr(target, key, value)
            changed = True
    return changed


def _extract_named_role(content: str, labels: tuple[str, ...]) -> tuple[str, str]:
    label_pattern = "|".join(re.escape(label) for label in labels)
    patterns = [
        re.compile(rf"(?:{label_pattern})\s*[:\-]?\s*(\d{{2}}기)\s*([가-힣]{{2,5}})"),
        re.compile(rf"(?:{label_pattern})\s*[:\-]?\s*([가-힣]{{2,5}})\s*\((\d{{2}}기)\)"),
        re.compile(rf"(?:{label_pattern})\s*[:\-]?\s*([가-힣]{{2,5}})[^\n]*\n\s*(\d{{2}}기)"),
    ]
    for index, pattern in enumerate(patterns):
        match = pattern.search(content)
        if not match:
            continue
        first, second = match.groups()
        if index == 0:
            return second, first
        return first, second
    return "", ""


def _cohort_leader_entry(row: SourceRow) -> dict[str, Any]:
    content = normalize_content(row.data.get("content"))
    author_name, cohort = normalize_author(row.data.get("writeUser"), row.data.get("cohort"))
    cohort = cohort or normalize_cohort(f"{value_as_text(row.data.get('title'))}\n{content}") or "기수 미상"
    captain, _ = _extract_named_role(content, ("기장",))
    vice, _ = _extract_named_role(content, ("부기장",))
    lines = [line.strip() for line in content.splitlines() if line.strip()]
    greeting = next((line for line in lines[1:] if "기장" not in line), lines[0] if lines else "")
    return {
        "cohort": cohort.removesuffix("기"),
        "captain_name": captain or author_name,
        "vice_captain_name": vice,
        "greeting": greeting[:500],
        "intro": redact_text(content)[0],
        "banner_image_url": "",
        "captain_image_url": "",
        "vice_captain_image_url": "",
        "legacy_write_id": row.source_id,
    }


def _past_council_entry(row: SourceRow) -> dict[str, Any]:
    title = value_as_text(row.data.get("title"))
    content = normalize_content(row.data.get("content"))
    council_match = re.search(r"(?:제\s*)?(\d{2})대", f"{title}\n{content}")
    president_name, president_cohort = _extract_named_role(content, ("원우회장", "회장"))
    vice_name, vice_cohort = _extract_named_role(content, ("원우부회장", "부원우회장", "부회장"))
    return {
        "cohort": council_match.group(1) if council_match else "미상",
        "president_name": president_name or "확인 필요",
        "president_cohort": president_cohort,
        "vice_president_name": vice_name,
        "vice_president_cohort": vice_cohort,
        "intro": redact_text(content)[0],
        "activities": [],
        "banner_image_url": "",
        "president_image_url": "",
        "vice_president_image_url": "",
        "legacy_write_id": row.source_id,
    }


def import_articles_and_specials(
    db: Session,
    rows: list[SourceRow],
    duplicates: list[SourceRow],
    *,
    apply: bool,
    limit: int | None = None,
) -> tuple[dict[str, Post], Counter, set[str]]:
    stats: Counter = Counter()
    boards = {board.slug: board for board in db.scalars(select(Board)).all()}
    if "free-board" in boards:
        raise RuntimeError("The forbidden free-board exists in the target database.")
    existing_posts = _existing_posts_by_legacy_id(db)
    user_cache = _legacy_user_cache(db)
    authoritative_ids = {row.source_id for row in rows}
    selected_rows = rows[:limit] if limit is not None else rows
    cohort_entries: list[dict[str, Any]] = []
    council_entries: list[dict[str, Any]] = []
    posts_by_source_id: dict[str, Post] = {}

    for duplicate in duplicates:
        stats["duplicate_rows"] += 1
        if apply:
            upsert_ledger(
                db,
                duplicate,
                entity_type="article",
                action="skipped",
                status="archived",
                reason="duplicate_article_id_in_authoritative_workbook",
                details={"title": redact_text(duplicate.data.get("title"))[0]},
            )

    for row in selected_rows:
        special_type = SPECIAL_SHEETS.get(row.sheet)
        existing_post = existing_posts.get(row.source_id)
        if special_type == "mutual_aid_archive":
            stats["archived_mutual_aid"] += 1
            if apply:
                if existing_post is not None and existing_post.status != "deleted":
                    existing_post.status = "deleted"
                    existing_post.deleted_at = utc_now()
                redacted_title, findings = redact_text(row.data.get("title"))
                upsert_ledger(
                    db,
                    row,
                    entity_type="mutual_aid_archive",
                    action="archived",
                    status="archived",
                    target_table="posts" if existing_post else None,
                    target_id=existing_post.id if existing_post else None,
                    reason="historical_sensitive_request_not_publicly_migrated",
                    details={"title": redacted_title, "pii_types": findings},
                )
            continue
        if special_type == "faq":
            question = shortened_title(redact_text(row.data.get("title"))[0], 500)
            answer = redact_text(normalize_content(row.data.get("content")))[0]
            if apply:
                ledger = db.scalar(
                    select(LegacyImportRecord).where(
                        LegacyImportRecord.source_file == row.source_file,
                        LegacyImportRecord.source_sheet == row.sheet,
                        LegacyImportRecord.entity_type == "faq",
                        LegacyImportRecord.source_id == row.source_id,
                    )
                )
                faq = db.get(FAQ, ledger.target_id) if ledger and ledger.target_table == "faqs" else None
                if faq is None:
                    faq = db.scalar(select(FAQ).where(FAQ.question == question))
                action = "unchanged"
                if faq is None:
                    faq = FAQ(question=question, answer=answer, category="legacy", sort_order=row.row_number)
                    db.add(faq)
                    db.flush()
                    action = "created"
                elif _assign_if_changed(
                    faq,
                    {"question": question, "answer": answer, "category": "legacy", "is_active": True},
                ):
                    action = "updated"
                if existing_post is not None and existing_post.status != "deleted":
                    existing_post.status = "deleted"
                    existing_post.deleted_at = utc_now()
                upsert_ledger(
                    db,
                    row,
                    entity_type="faq",
                    action=action,
                    status="imported",
                    target_table="faqs",
                    target_id=faq.id,
                )
            stats["faqs"] += 1
            continue
        if special_type == "cohort_leader":
            cohort_entries.append(_cohort_leader_entry(row))
            stats["cohort_leaders"] += 1
            if apply and existing_post is not None and existing_post.status != "deleted":
                existing_post.status = "deleted"
                existing_post.deleted_at = utc_now()
            continue
        if special_type == "past_council":
            council_entries.append(_past_council_entry(row))
            stats["past_councils"] += 1
            if apply and existing_post is not None and existing_post.status != "deleted":
                existing_post.status = "deleted"
                existing_post.deleted_at = utc_now()
            continue

        slug = target_slug(row)
        if slug is None:
            stats["unmapped_articles"] += 1
            if apply:
                upsert_ledger(
                    db,
                    row,
                    entity_type="article",
                    action="review",
                    status="unmapped",
                    reason="no_approved_board_mapping",
                    details={
                        "board_name": value_as_text(row.data.get("boardName")),
                        "title": redact_text(row.data.get("title"))[0],
                    },
                )
            continue
        board = boards.get(slug)
        if board is None:
            stats["missing_target_board"] += 1
            if apply:
                upsert_ledger(
                    db,
                    row,
                    entity_type="article",
                    action="failed",
                    status="failed",
                    reason=f"missing_target_board:{slug}",
                )
            continue
        if not apply:
            stats[f"mapped:{slug}"] += 1
            if existing_post is not None:
                posts_by_source_id[row.source_id] = existing_post
            else:
                posts_by_source_id[row.source_id] = Post(
                    id=-value_as_int(row.source_id, len(posts_by_source_id) + 1),
                    board_id=board.id,
                    author_id=None,
                    title=value_as_text(row.data.get("title"))[:200] or "제목 없음",
                    content=value_as_text(row.data.get("content")) or "검토 예정",
                    created_at=parse_datetime(row.data.get("date")) or datetime(1970, 1, 1),
                    updated_at=parse_datetime(row.data.get("updateDate"))
                    or parse_datetime(row.data.get("date"))
                    or datetime(1970, 1, 1),
                )
            continue
        preferred_user = db.get(User, existing_post.author_id) if existing_post and existing_post.author_id else None
        author = ensure_legacy_user(
            db,
            row.data.get("writeUser"),
            row.data.get("cohort"),
            user_cache,
            preferred=preferred_user,
        )
        try:
            source_legacy_id = (
                LEGACY_STUDY_ACTIVITY_SOURCE_WRITE_IDS.get(row.source_id)
                if board.slug == "study-activity"
                else None
            )
            source_post_candidate = (
                posts_by_source_id.get(source_legacy_id)
                or existing_posts.get(source_legacy_id)
                if source_legacy_id is not None
                else None
            )
            study_recruit_board = boards.get("study-recruit")
            source_post = (
                source_post_candidate
                if source_post_candidate is not None
                and study_recruit_board is not None
                and source_post_candidate.board_id == study_recruit_board.id
                else None
            )
            payload = _post_payload(
                row,
                board,
                author,
                activity_source_post_id=source_post.id if source_post is not None else None,
            )
        except ValueError as exc:
            stats["failed_articles"] += 1
            upsert_ledger(
                db,
                row,
                entity_type="article",
                action="failed",
                status="failed",
                reason=str(exc),
                details={"title": redact_text(row.data.get("title"))[0]},
            )
            continue
        if existing_post is not None:
            # A resumed import must not undo engagement accumulated after the
            # first migration rehearsal or after launch.
            payload["view_count"] = max(existing_post.view_count, payload["view_count"])
            payload["like_count"] = max(existing_post.like_count, payload["like_count"])
        if existing_post is None:
            post = Post(comment_count=0, **payload)
            db.add(post)
            db.flush()
            existing_posts[row.source_id] = post
            action = "created"
            stats["created_posts"] += 1
        else:
            post = existing_post
            changed_fields = [key for key, value in payload.items() if getattr(post, key) != value]
            changed = _assign_if_changed(post, payload)
            action = "updated" if changed else "unchanged"
            stats[f"{action}_posts"] += 1
            for field in changed_fields:
                stats[f"updated_post_field:{field}"] += 1
        posts_by_source_id[row.source_id] = post
        if board.board_type == "suggestion":
            suggestion = db.scalar(select(PostSuggestion).where(PostSuggestion.post_id == post.id))
            if suggestion is None:
                db.add(PostSuggestion(post_id=post.id, suggestion_category="legacy", status="received"))
        upsert_ledger(
            db,
            row,
            entity_type="article",
            action=action,
            status="imported",
            target_table="posts",
            target_id=post.id,
            details={"board_slug": slug, "title": post.title},
        )

    if apply:
        study_activity_board = boards.get("study-activity")
        study_recruit_board = boards.get("study-recruit")
        for activity_legacy_id, source_legacy_id in LEGACY_STUDY_ACTIVITY_SOURCE_WRITE_IDS.items():
            activity_post = posts_by_source_id.get(activity_legacy_id)
            source_post = posts_by_source_id.get(source_legacy_id) or existing_posts.get(source_legacy_id)
            if (
                activity_post is None
                or source_post is None
                or study_activity_board is None
                or study_recruit_board is None
                or activity_post.board_id != study_activity_board.id
                or source_post.board_id != study_recruit_board.id
            ):
                continue
            metadata = dict(activity_post.metadata_json or {})
            if metadata.get("activity_source_post_id") != str(source_post.id):
                metadata["activity_source_post_id"] = str(source_post.id)
                activity_post.metadata_json = metadata

        for special_rows, board_slug, metadata_key, entity_type in (
            (cohort_entries, "gsa-cohort-leaders", "cohort_leaders", "cohort_leader"),
            (council_entries, "gsa-past-councils", "past_councils", "past_council"),
        ):
            if not special_rows:
                continue
            board = boards.get(board_slug)
            if board is None:
                raise RuntimeError(f"Missing required special board: {board_slug}")
            metadata = copy.deepcopy(board.metadata_json or {})
            existing_entries = {
                str(entry.get("legacy_write_id")): entry
                for entry in metadata.get(metadata_key, [])
                if entry.get("legacy_write_id") is not None
            }
            for entry in special_rows:
                existing_entry = existing_entries.get(str(entry["legacy_write_id"]))
                if existing_entry is None:
                    continue
                for field in ("banner_image_url", "attachment_urls"):
                    if existing_entry.get(field):
                        entry[field] = copy.deepcopy(existing_entry[field])
            metadata[metadata_key] = special_rows
            metadata_changed = board.metadata_json != metadata
            if metadata_changed:
                board.metadata_json = metadata
            row_by_id = {row.source_id: row for row in selected_rows if SPECIAL_SHEETS.get(row.sheet) == entity_type}
            for entry in special_rows:
                source_id = str(entry["legacy_write_id"])
                upsert_ledger(
                    db,
                    row_by_id[source_id],
                    entity_type=entity_type,
                    action="updated" if metadata_changed else "unchanged",
                    status="imported",
                    target_table="boards",
                    target_id=board.id,
                    details={"board_slug": board_slug, "cohort": entry["cohort"]},
                )

        if limit is None:
            for source_id, post in existing_posts.items():
                if source_id in authoritative_ids or post.status == "deleted":
                    continue
                post.status = "deleted"
                post.deleted_at = utc_now()
                synthetic = SourceRow(
                    "board_articles_ver2.xlsx",
                    value_as_text((post.metadata_json or {}).get("legacy_sheet")) or "legacy_db_only",
                    0,
                    {"writeId": source_id, "title": post.title, "metadata": post.metadata_json or {}},
                )
                upsert_ledger(
                    db,
                    synthetic,
                    entity_type="obsolete_article",
                    action="soft_deleted",
                    status="archived",
                    target_table="posts",
                    target_id=post.id,
                    reason="absent_from_authoritative_ver2",
                    details={"title": redact_text(post.title)[0]},
                )
                stats["soft_deleted_obsolete_posts"] += 1
    return posts_by_source_id, stats, authoritative_ids


def import_comments(
    db: Session,
    rows: list[SourceRow],
    posts_by_source_id: dict[str, Post],
    *,
    apply: bool,
    reconcile_untracked: bool = True,
) -> Counter:
    stats: Counter = Counter()
    board_slug_by_id = {board.id: board.slug for board in db.scalars(select(Board)).all()}
    if apply and reconcile_untracked:
        tracked_comment_ids = {
            target_id
            for (target_id,) in db.execute(
                select(LegacyImportRecord.target_id).where(
                    LegacyImportRecord.entity_type == "comment",
                    LegacyImportRecord.status == "imported",
                    LegacyImportRecord.target_table == "comments",
                    LegacyImportRecord.target_id.is_not(None),
                )
            ).all()
        }
        users = {user.id: user for user in db.scalars(select(User)).all()}
        posts = {post.id: post for post in db.scalars(select(Post)).all()}
        for existing_comment in db.scalars(select(Comment)).all():
            if existing_comment.id in tracked_comment_ids:
                continue
            author = users.get(existing_comment.author_id or -1)
            post = posts.get(existing_comment.post_id)
            metadata = post.metadata_json if post else None
            if (
                author is not None
                and author.username.startswith("legacy")
                and isinstance(metadata, dict)
                and (metadata.get("legacy_write_id") is not None or metadata.get("legacy_article_id") is not None)
            ):
                db.delete(existing_comment)
                stats["removed_obsolete_legacy_comments"] += 1
        if stats["removed_obsolete_legacy_comments"]:
            db.flush()
    existing_by_key: dict[tuple[int, datetime, str], list[Comment]] = defaultdict(list)
    for comment in db.scalars(select(Comment)).all():
        existing_by_key[(comment.post_id, comment.created_at, comment.content)].append(comment)
    user_cache = _legacy_user_cache(db)
    comment_by_source_id: dict[str, Comment] = {}

    for pass_replies in (False, True):
        for row in rows:
            parent_source_id = value_as_text(row.data.get("parent_comment_id"))
            if bool(parent_source_id) != pass_replies:
                continue
            article_id = value_as_text(row.data.get("article_id"))
            post = posts_by_source_id.get(article_id)
            if post is None:
                stats["orphan_comments"] += 1
                if apply:
                    upsert_ledger(
                        db,
                        row,
                        entity_type="comment",
                        action="archived",
                        status="archived",
                        source_parent_id=parent_source_id or None,
                        reason="article_not_publicly_migrated",
                        details={"article_id": article_id, "preview": redact_text(row.data.get("content"))[0][:160]},
                    )
                continue
            slug = board_slug_by_id.get(post.board_id, "")
            if slug in COMMENT_DISABLED_SLUGS:
                stats["policy_archived_comments"] += 1
                if apply:
                    upsert_ledger(
                        db,
                        row,
                        entity_type="comment",
                        action="archived",
                        status="archived",
                        source_parent_id=parent_source_id or None,
                        target_table="posts",
                        target_id=post.id,
                        reason=f"comments_not_migrated_for_board:{slug}",
                        details={"preview": redact_text(row.data.get("content"))[0][:160]},
                    )
                continue
            parent = comment_by_source_id.get(parent_source_id) if parent_source_id else None
            if parent_source_id and parent is None:
                stats["orphan_reply_comments"] += 1
                if apply:
                    upsert_ledger(
                        db,
                        row,
                        entity_type="comment",
                        action="archived",
                        status="archived",
                        source_parent_id=parent_source_id,
                        reason="missing_or_archived_parent_comment",
                        details={"preview": redact_text(row.data.get("content"))[0][:160]},
                    )
                continue
            if parent is not None and parent.parent_id is not None:
                stats["over_depth_reply_comments"] += 1
                if apply:
                    upsert_ledger(
                        db,
                        row,
                        entity_type="comment",
                        action="archived",
                        status="archived",
                        source_parent_id=parent_source_id,
                        reason="comment_depth_exceeds_two",
                        details={"preview": redact_text(row.data.get("content"))[0][:160]},
                    )
                continue
            original_content = value_as_text(row.data.get("content"))
            content, findings = redact_text(original_content)
            created_at = parse_datetime(row.data.get("date"))
            if not content or created_at is None:
                stats["failed_comments"] += 1
                continue
            if not apply:
                stats["mapped_comments"] += 1
                comment_by_source_id[row.source_id] = Comment(
                    id=-value_as_int(row.source_id, len(comment_by_source_id) + 1),
                    post_id=post.id,
                    author_id=None,
                    parent_id=parent.id if parent else None,
                    content=content,
                    created_at=created_at,
                    updated_at=parse_datetime(row.data.get("update_date")) or created_at,
                )
                continue
            candidates = existing_by_key.get((post.id, created_at, content), [])
            if not candidates and original_content != content:
                candidates = existing_by_key.get((post.id, created_at, original_content), [])
            comment = candidates.pop(0) if candidates else None
            action = "unchanged"
            author = ensure_legacy_user(
                db,
                row.data.get("write_user"),
                row.data.get("cohort"),
                user_cache,
                preferred=db.get(User, comment.author_id) if comment and comment.author_id else None,
            )
            if comment is None:
                comment = Comment(
                    post_id=post.id,
                    author_id=author.id,
                    author_nickname_snapshot=author.nickname,
                    author_cohort_snapshot=author.cohort,
                    parent_id=parent.id if parent else None,
                    content=content,
                    created_at=created_at,
                    updated_at=parse_datetime(row.data.get("update_date")) or created_at,
                )
                db.add(comment)
                db.flush()
                action = "created"
            else:
                desired_parent_id = parent.id if parent else None
                if comment.content != content:
                    comment.content = content
                    action = "updated"
                if comment.parent_id != desired_parent_id:
                    comment.parent_id = desired_parent_id
                    action = "updated"
                if comment.author_nickname_snapshot is None:
                    comment.author_nickname_snapshot = author.nickname
                    action = "updated"
                if comment.author_cohort_snapshot is None:
                    comment.author_cohort_snapshot = author.cohort
                    action = "updated"
            comment_by_source_id[row.source_id] = comment
            upsert_ledger(
                db,
                row,
                entity_type="comment",
                action=action,
                status="imported",
                source_parent_id=parent_source_id or None,
                target_table="comments",
                target_id=comment.id,
                reason="pii_redacted" if findings else None,
                details={"article_id": article_id, "pii_types": findings},
            )
            stats[f"{action}_comments"] += 1

    if apply:
        for post in posts_by_source_id.values():
            actual_count = int(db.scalar(select(func.count(Comment.id)).where(Comment.post_id == post.id)) or 0)
            post.comment_count = actual_count
            legacy_updated_at = value_as_text((post.metadata_json or {}).get("legacy_updated_at"))
            restored_updated_at = parse_datetime(legacy_updated_at)
            if restored_updated_at is not None and db.is_modified(post, include_collections=False):
                post.updated_at = restored_updated_at
                # Force an explicit value so SQLAlchemy's on-update default cannot
                # replace the historical timestamp when another field changed.
                flag_modified(post, "updated_at")
    return stats


def _safe_filename(value: str, fallback: str) -> str:
    basename = Path(value.replace("\\", "/")).name.strip()
    basename = re.sub(r"[\x00-\x1f<>:\"/\\|?*]", "_", basename)
    return (basename or fallback)[:255]


def _legacy_attachment_filename(subject: object, storage_id: str, url: str) -> str:
    submitted_name = _safe_filename(value_as_text(subject), f"legacy-{storage_id}")
    if Path(submitted_name).suffix:
        return submitted_name
    url_name = _safe_filename(
        urllib.parse.unquote(Path(urllib.parse.urlparse(url).path).name),
        "",
    )
    if Path(url_name).suffix:
        return url_name
    return submitted_name


def _download_url(value: str) -> str:
    """Quote legacy object-storage paths without changing query parameters."""
    parsed = urllib.parse.urlsplit(value)
    quoted_path = urllib.parse.quote(urllib.parse.unquote(parsed.path), safe="/:@")
    return urllib.parse.urlunsplit(
        (parsed.scheme, parsed.netloc, quoted_path, parsed.query, parsed.fragment)
    )


def _detected_content_type(path: Path, *candidates: str) -> str:
    """Return the MIME type supported by the app that matches the file bytes.

    Legacy storage headers and filenames are not reliable. Candidate types still
    take precedence for OLE containers, whose magic bytes alone cannot distinguish
    Word, Excel, PowerPoint, and HWP documents.
    """
    detection_order = [
        *(normalize_content_type(candidate) for candidate in candidates),
        "image/jpeg",
        "image/png",
        "image/gif",
        "image/webp",
        "image/heic",
        "image/heif",
        "application/pdf",
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.openxmlformats-officedocument.presentationml.presentation",
        "application/zip",
        "application/x-ipynb+json",
        "application/json",
        "text/plain",
        "application/msword",
        "application/vnd.ms-excel",
        "application/vnd.ms-powerpoint",
        "application/x-hwp",
        "application/haansofthwp",
        "application/vnd.hancom.hwp",
    ]
    seen: set[str] = set()
    for content_type in detection_order:
        if content_type in seen or content_type not in ALLOWED_MIME_TYPES:
            continue
        seen.add(content_type)
        if _matches_declared_content(path, content_type):
            return content_type
    return ""


def _attach_media_to_special_entry(
    board: Board,
    *,
    collection_key: str,
    article_id: str,
    media_url: str,
) -> bool:
    metadata = copy.deepcopy(board.metadata_json or {})
    entries = list(metadata.get(collection_key) or [])
    changed = False
    for entry in entries:
        if str(entry.get("legacy_write_id")) != article_id:
            continue
        attachment_urls = list(entry.get("attachment_urls") or [])
        if media_url not in attachment_urls:
            attachment_urls.append(media_url)
            entry["attachment_urls"] = attachment_urls
            changed = True
        if not entry.get("banner_image_url"):
            entry["banner_image_url"] = media_url
            changed = True
        break
    if changed:
        metadata[collection_key] = entries
        board.metadata_json = metadata
    return changed


def _download_attachment(
    url: str,
    destination: Path,
    *,
    declared_name: str,
    maximum_bytes: int,
) -> tuple[str, int, str, Path]:
    request = urllib.request.Request(
        _download_url(url),
        headers={"User-Agent": "AISW-Legacy-Migration/1.0"},
    )
    temporary = destination.with_name(f".{destination.name}.part")
    temporary.parent.mkdir(parents=True, exist_ok=True)
    temporary.unlink(missing_ok=True)
    digest = hashlib.sha256()
    total = 0
    try:
        with urllib.request.urlopen(request, timeout=45) as response, temporary.open("xb") as output:
            content_type = normalize_content_type(response.headers.get_content_type())
            declared_length = response.headers.get("Content-Length")
            if declared_length and int(declared_length) > maximum_bytes:
                raise ValueError("file_too_large")
            while True:
                chunk = response.read(1024 * 1024)
                if not chunk:
                    break
                total += len(chunk)
                if total > maximum_bytes:
                    raise ValueError("file_too_large")
                digest.update(chunk)
                output.write(chunk)
        if total == 0:
            raise ValueError("empty_file")
        inferred_content_type = normalize_content_type(mimetypes.guess_type(declared_name)[0])
        actual_content_type = _detected_content_type(
            temporary,
            content_type,
            inferred_content_type,
        )
        if not actual_content_type:
            reported_type = content_type or inferred_content_type or "unknown"
            raise ValueError(f"unsupported_or_invalid_content:{reported_type}")
        expected_extension = MIME_EXTENSIONS[actual_content_type]
        corrected_destination = destination.with_suffix(expected_extension)
        corrected_destination.parent.mkdir(parents=True, exist_ok=True)
        os.replace(temporary, corrected_destination)
        return actual_content_type, total, digest.hexdigest(), corrected_destination
    except BaseException:
        temporary.unlink(missing_ok=True)
        raise


def index_local_attachment_files(source_dir: Path) -> dict[str, Path]:
    """Index migration input files by the legacy ``fileStorageId`` basename."""

    source_dir = source_dir.expanduser().resolve()
    if not source_dir.is_dir():
        raise ValueError(f"attachment_source_directory_not_found:{source_dir}")
    result: dict[str, Path] = {}
    duplicate_ids: set[str] = set()
    for path in sorted(source_dir.rglob("*")):
        if not path.is_file():
            continue
        if path.is_symlink():
            raise ValueError(f"symlinked_attachment_source_not_allowed:{path.name}")
        storage_id = path.stem.strip()
        if not storage_id:
            continue
        if storage_id in result:
            duplicate_ids.add(storage_id)
            continue
        result[storage_id] = path
    if duplicate_ids:
        rendered = ",".join(sorted(duplicate_ids)[:20])
        raise ValueError(f"duplicate_local_attachment_storage_ids:{rendered}")
    return result


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        while chunk := handle.read(1024 * 1024):
            digest.update(chunk)
    return digest.hexdigest()


def _copy_local_attachment(
    source: Path,
    destination: Path,
    *,
    declared_name: str,
    maximum_bytes: int,
) -> tuple[str, int, str, Path]:
    """Validate and atomically copy a local migration file into managed media storage."""

    if not source.is_file() or source.is_symlink():
        raise ValueError("invalid_local_attachment_source")
    temporary = destination.with_name(f".{destination.name}.part")
    temporary.parent.mkdir(parents=True, exist_ok=True)
    temporary.unlink(missing_ok=True)
    digest = hashlib.sha256()
    total = 0
    try:
        with source.open("rb") as input_file, temporary.open("xb") as output_file:
            while chunk := input_file.read(1024 * 1024):
                total += len(chunk)
                if total > maximum_bytes:
                    raise ValueError("file_too_large")
                digest.update(chunk)
                output_file.write(chunk)
        if total == 0:
            raise ValueError("empty_file")
        inferred_content_type = normalize_content_type(
            mimetypes.guess_type(declared_name)[0] or mimetypes.guess_type(source.name)[0]
        )
        actual_content_type = _detected_content_type(temporary, inferred_content_type)
        if not actual_content_type:
            raise ValueError(
                f"unsupported_or_invalid_content:{inferred_content_type or 'unknown'}"
            )
        corrected_destination = destination.with_suffix(MIME_EXTENSIONS[actual_content_type])
        corrected_destination.parent.mkdir(parents=True, exist_ok=True)
        os.replace(temporary, corrected_destination)
        return actual_content_type, total, digest.hexdigest(), corrected_destination
    except BaseException:
        temporary.unlink(missing_ok=True)
        raise


def import_attachments(
    db: Session,
    rows: list[SourceRow],
    reference_urls: dict[str, str],
    posts_by_source_id: dict[str, Post],
    *,
    media_root: Path | None = None,
    public_media_dir: Path | None = None,
    private_media_dir: Path | None = None,
    attachment_source_dir: Path | None = None,
    apply: bool,
    skip_downloads: bool,
    maximum_bytes: int = 20 * 1024 * 1024,
) -> Counter:
    stats: Counter = Counter()
    if (public_media_dir is None) != (private_media_dir is None):
        raise ValueError("public_and_private_media_directories_must_be_set_together")
    if public_media_dir is not None and private_media_dir is not None:
        public_root = public_media_dir.expanduser().resolve()
        private_root = private_media_dir.expanduser().resolve()
    elif media_root is not None:
        resolved_media_root = media_root.expanduser().resolve()
        public_root = resolved_media_root / "public"
        private_root = resolved_media_root / "private"
    else:
        raise ValueError("media_storage_directories_are_required")
    if public_root == private_root:
        raise ValueError("public_and_private_media_directories_must_be_distinct")
    local_files = (
        index_local_attachment_files(attachment_source_dir)
        if attachment_source_dir is not None
        else {}
    )
    existing_media = {media.stored_filename: media for media in db.scalars(select(MediaAsset)).all()}
    existing_media_by_storage_id = {
        match.group(1): media
        for media in existing_media.values()
        if (match := re.fullmatch(r"legacy-(.+?)(?:\.[^.]+)?", media.stored_filename))
    }
    ledger_by_article = {
        (record.entity_type, record.source_id): record
        for record in db.scalars(
            select(LegacyImportRecord).where(
                LegacyImportRecord.entity_type.in_(["cohort_leader", "past_council", "mutual_aid_archive", "faq"])
            )
        ).all()
    }
    board_by_id = {board.id: board for board in db.scalars(select(Board)).all()}
    media_hashes_by_post: dict[int, dict[str, int]] = defaultdict(dict)

    for index, row in enumerate(rows, start=1):
        if apply and not skip_downloads and (index == 1 or index % 25 == 0 or index == len(rows)):
            print(f"attachment progress: {index}/{len(rows)}", flush=True)
        article_id = value_as_text(row.data.get("writeId"))
        storage_id = value_as_text(row.data.get("fileStorageId"))
        if not article_id or not storage_id:
            stats["invalid_attachment_rows"] += 1
            continue
        url = value_as_text(row.data.get("link_url")) or reference_urls.get(storage_id, "")
        local_source = local_files.get(storage_id)
        source_id = storage_id
        attachment_row = SourceRow(row.source_file, row.sheet, row.row_number, {**row.data, "writeId": source_id})
        post = posts_by_source_id.get(article_id)
        special_record = next(
            (
                ledger_by_article[key]
                for key in (
                    ("cohort_leader", article_id),
                    ("past_council", article_id),
                    ("mutual_aid_archive", article_id),
                    ("faq", article_id),
                )
                if key in ledger_by_article
            ),
            None,
        )
        is_private = bool(
            (special_record and special_record.entity_type == "mutual_aid_archive")
            or "/private/" in url.lower()
        )
        if post is None and special_record is None:
            stats["orphan_attachments"] += 1
            if apply:
                upsert_ledger(
                    db,
                    attachment_row,
                    entity_type="attachment",
                    action="archived",
                    status="archived",
                    source_parent_id=article_id,
                    reason="article_not_migrated",
                    details={"article_id": article_id},
                )
            continue
        if attachment_source_dir is not None and local_source is None:
            stats["missing_local_attachment_files"] += 1
            if apply:
                upsert_ledger(
                    db,
                    attachment_row,
                    entity_type="attachment",
                    action="failed",
                    status="failed",
                    source_parent_id=article_id,
                    reason="missing_local_attachment_file",
                    details={"article_id": article_id, "filename": value_as_text(row.data.get("subject"))},
                )
            continue
        if local_source is None and not url.startswith(("http://", "https://")):
            stats["missing_attachment_urls"] += 1
            if apply:
                upsert_ledger(
                    db,
                    attachment_row,
                    entity_type="attachment",
                    action="failed",
                    status="failed",
                    source_parent_id=article_id,
                    reason="missing_download_url",
                    details={"article_id": article_id, "filename": value_as_text(row.data.get("subject"))},
                )
            continue
        filename = _legacy_attachment_filename(row.data.get("subject"), storage_id, url)
        guessed_extension = (
            Path(filename).suffix.lower()
            or (local_source.suffix.lower() if local_source is not None else "")
            or Path(urllib.parse.urlparse(url).path).suffix.lower()
        )
        if local_source is not None and local_source.suffix.lower() in ARCHIVED_LEGACY_ATTACHMENT_EXTENSIONS:
            stats["archived_unsupported_attachments"] += 1
            if apply:
                upsert_ledger(
                    db,
                    attachment_row,
                    entity_type="attachment",
                    action="archived",
                    status="archived",
                    source_parent_id=article_id,
                    reason="legacy_attachment_type_not_supported",
                    details={
                        "article_id": article_id,
                        "filename": filename,
                        "extension": local_source.suffix.lower(),
                        "source_origin": "local",
                    },
                )
            continue
        stored_filename = f"legacy-{storage_id}{guessed_extension}"
        destination = (private_root if is_private else public_root) / stored_filename
        media = existing_media.get(stored_filename) or existing_media_by_storage_id.get(storage_id)
        if media is not None:
            stored_filename = media.stored_filename
            destination = (private_root if media.is_private else public_root) / stored_filename
        if skip_downloads or not apply:
            stats["local_copy_candidates" if local_source is not None else "download_candidates"] += 1
            continue
        action = "unchanged"
        sha256 = ""
        stored_now = False
        renamed_from: Path | None = None
        post_attachment_id: int | None = None
        faq_attachment_id: int | None = None
        duplicate_of_media_asset_id: int | None = None
        try:
            if destination.exists() and destination.stat().st_size > 0:
                sha256 = _sha256_file(destination)
                content_type = _detected_content_type(
                    destination,
                    media.content_type if media else "",
                    mimetypes.guess_type(filename)[0] or "",
                )
                if not content_type:
                    raise ValueError("unsupported_or_invalid_existing_file")
                file_size = destination.stat().st_size
                corrected_destination = destination.with_suffix(MIME_EXTENSIONS[content_type])
                if corrected_destination != destination:
                    if corrected_destination.exists():
                        raise ValueError("corrected_attachment_destination_exists")
                    renamed_from = destination
                    os.replace(destination, corrected_destination)
                    destination = corrected_destination
                stored_filename = destination.name
            elif local_source is not None:
                content_type, file_size, sha256, destination = _copy_local_attachment(
                    local_source,
                    destination,
                    declared_name=filename,
                    maximum_bytes=maximum_bytes,
                )
                stored_now = True
                stored_filename = destination.name
            else:
                content_type, file_size, sha256, destination = _download_attachment(
                    url,
                    destination,
                    declared_name=filename,
                    maximum_bytes=maximum_bytes,
                )
                stored_now = True
                stored_filename = destination.name
            if media is None:
                media = MediaAsset(
                    owner_id=post.author_id if post else None,
                    original_filename=filename,
                    stored_filename=stored_filename,
                    content_type=content_type,
                    file_size=file_size,
                    url=None,
                    is_private=is_private,
                    status="ready",
                    created_at=parse_datetime(row.data.get("regiDatetime")) or utc_now(),
                )
                db.add(media)
                db.flush()
                media.url = media_access_reference(media.id)
                existing_media[stored_filename] = media
                existing_media_by_storage_id[storage_id] = media
                action = "created"
            else:
                changed = _assign_if_changed(
                    media,
                    {
                        "stored_filename": stored_filename,
                        "original_filename": filename,
                        "content_type": content_type,
                        "file_size": file_size,
                        "url": media_access_reference(media.id),
                        "is_private": is_private,
                        "status": "ready",
                    },
                )
                action = "updated" if changed else "unchanged"
            if post is not None:
                duplicate_of_media_asset_id = media_hashes_by_post[post.id].get(sha256)
                link = db.scalar(
                    select(PostAttachment).where(
                        PostAttachment.post_id == post.id,
                        PostAttachment.media_id == media.id,
                    )
                )
                if duplicate_of_media_asset_id is not None and duplicate_of_media_asset_id != media.id:
                    if link is not None:
                        db.delete(link)
                        db.flush()
                    stats["deduplicated_post_attachment_links"] += 1
                else:
                    media_hashes_by_post[post.id][sha256] = media.id
                    sequence = value_as_int(row.data.get("sequence"))
                    if link is None:
                        link = PostAttachment(
                            post_id=post.id,
                            media_id=media.id,
                            sort_order=sequence,
                        )
                        db.add(link)
                        db.flush()
                    elif link.sort_order != sequence:
                        link.sort_order = sequence
                    post_attachment_id = link.id
            elif special_record and special_record.entity_type in {"cohort_leader", "past_council"}:
                board = board_by_id.get(special_record.target_id or 0)
                if board:
                    collection_key = "cohort_leaders" if special_record.entity_type == "cohort_leader" else "past_councils"
                    _attach_media_to_special_entry(
                        board,
                        collection_key=collection_key,
                        article_id=article_id,
                        media_url=media_access_reference(media.id),
                    )
            elif special_record and special_record.entity_type == "faq":
                faq = db.get(FAQ, special_record.target_id or 0)
                if faq is not None:
                    faq_link = db.scalar(
                        select(FAQAttachment).where(
                            FAQAttachment.faq_id == faq.id,
                            FAQAttachment.media_id == media.id,
                        )
                    )
                    sequence = value_as_int(row.data.get("sequence"))
                    if faq_link is None:
                        faq_link = FAQAttachment(
                            faq_id=faq.id,
                            media_id=media.id,
                            sort_order=sequence,
                        )
                        db.add(faq_link)
                        db.flush()
                    elif faq_link.sort_order != sequence:
                        faq_link.sort_order = sequence
                    faq_attachment_id = faq_link.id
            upsert_ledger(
                db,
                attachment_row,
                entity_type="attachment",
                action=action,
                status="imported",
                source_parent_id=article_id,
                target_table="media_assets",
                target_id=media.id,
                details={
                    "article_id": article_id,
                    "filename": filename,
                    "content_type": content_type,
                    "file_size": file_size,
                    "sha256": sha256,
                    "is_private": is_private,
                    "storage_path": f"{'private' if is_private else 'public'}/{stored_filename}",
                    "source_origin": "local" if local_source is not None else "download",
                    "source_url": None if local_source is not None else url,
                    "media_asset_id": media.id,
                    "post_attachment_id": post_attachment_id,
                    "faq_attachment_id": faq_attachment_id,
                    "duplicate_of_media_asset_id": duplicate_of_media_asset_id,
                },
            )
            stats[f"{action}_attachments"] += 1
        except Exception as exc:
            if stored_now:
                destination.unlink(missing_ok=True)
            elif renamed_from is not None and destination.exists():
                os.replace(destination, renamed_from)
            stats["failed_attachments"] += 1
            upsert_ledger(
                db,
                attachment_row,
                entity_type="attachment",
                action="failed",
                status="failed",
                source_parent_id=article_id,
                reason=str(exc)[:500],
                details={
                    "article_id": article_id,
                    "filename": filename,
                    "source_origin": "local" if local_source is not None else "download",
                    "source_url": None if local_source is not None else url,
                },
            )
    return stats


def summarize_sources(
    articles: list[SourceRow],
    comments: list[SourceRow],
    attachments: list[SourceRow],
    duplicates: list[SourceRow],
) -> dict[str, Any]:
    sheet_counts = Counter(row.sheet for row in articles)
    board_counts = Counter(target_slug(row) or SPECIAL_SHEETS.get(row.sheet, "UNMAPPED") for row in articles)
    staging_counts = Counter(
        value_as_text(row.data.get("boardName")) for row in articles if row.sheet == "자유게시판"
    )
    return {
        "articles": len(articles),
        "comments": len(comments),
        "attachments": len(attachments),
        "duplicates": len(duplicates),
        "sheet_counts": dict(sorted(sheet_counts.items())),
        "target_counts": dict(sorted(board_counts.items())),
        "staging_board_name_counts": dict(sorted(staging_counts.items())),
    }


def export_ledger_report(db: Session, report_dir: Path, summary: dict[str, Any]) -> tuple[Path, Path]:
    report_dir.mkdir(parents=True, exist_ok=True)
    records = db.scalars(select(LegacyImportRecord).order_by(LegacyImportRecord.id)).all()
    status_counts = Counter(record.status for record in records)
    entity_counts = Counter(record.entity_type for record in records)
    summary = {
        **summary,
        "ledger_records": len(records),
        "ledger_status_counts": dict(sorted(status_counts.items())),
        "ledger_entity_counts": dict(sorted(entity_counts.items())),
        "generated_at": utc_now().isoformat(),
    }
    summary_path = report_dir / "migration-summary.json"
    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    records_path = report_dir / "migration-records.ndjson"
    with records_path.open("w", encoding="utf-8") as handle:
        for record in records:
            handle.write(
                json.dumps(
                    {
                        "id": record.id,
                        "source_file": record.source_file,
                        "source_sheet": record.source_sheet,
                        "source_row": record.source_row,
                        "entity_type": record.entity_type,
                        "source_id": record.source_id,
                        "source_parent_id": record.source_parent_id,
                        "source_hash": record.source_hash,
                        "action": record.action,
                        "status": record.status,
                        "target_table": record.target_table,
                        "target_id": record.target_id,
                        "reason": record.reason,
                        "redacted_details": record.redacted_details,
                    },
                    ensure_ascii=False,
                    default=str,
                )
                + "\n"
            )
    return summary_path, records_path
