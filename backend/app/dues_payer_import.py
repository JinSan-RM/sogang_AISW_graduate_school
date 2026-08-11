from dataclasses import dataclass
from io import BytesIO
import re

from openpyxl import load_workbook

from app.errors import AppException


STUDENT_NUMBER_PATTERN = re.compile(r"^A\d{5}$")


@dataclass(frozen=True)
class DuesPayerRow:
    row_number: int
    name: str
    major: str
    student_number: str


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _workbook_error(message: str) -> AppException:
    return AppException(status_code=422, message=message, code="INVALID_DUES_WORKBOOK")


def parse_dues_payer_workbook(content: bytes) -> list[DuesPayerRow]:
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise _workbook_error("The uploaded file is not a readable .xlsx workbook.") from exc

    rows: list[DuesPayerRow] = []
    seen_student_numbers: set[str] = set()
    try:
        if not workbook.worksheets:
            raise _workbook_error("The workbook does not contain a worksheet.")
        sheet = workbook.worksheets[0]
        for row_number, values in enumerate(sheet.iter_rows(values_only=True), start=1):
            cells = list(values)
            if any(_cell_text(value) for value in cells[3:]):
                raise _workbook_error(f"Row {row_number} contains data outside the name, major, and student-number columns.")

            cells.extend([None] * (3 - len(cells)))
            name, major, student_number = (_cell_text(value) for value in cells[:3])
            if not name and not major and not student_number:
                continue
            if not name or not major or not student_number:
                raise AppException(
                    status_code=422,
                    message=f"Row {row_number} has an empty name, major, or student number.",
                    code="DUES_IMPORT_EMPTY_VALUE",
                )

            student_number = student_number.upper()
            if not STUDENT_NUMBER_PATTERN.fullmatch(student_number):
                raise AppException(
                    status_code=422,
                    message=f"Row {row_number} has an invalid student number.",
                    code="DUES_IMPORT_INVALID_STUDENT_NUMBER",
                )
            if student_number in seen_student_numbers:
                raise AppException(
                    status_code=422,
                    message=f"Row {row_number} duplicates student number {student_number}.",
                    code="DUES_IMPORT_DUPLICATE_STUDENT_NUMBER",
                )
            seen_student_numbers.add(student_number)
            rows.append(
                DuesPayerRow(
                    row_number=row_number,
                    name=name,
                    major=major,
                    student_number=student_number,
                )
            )
    finally:
        workbook.close()

    if not rows:
        raise _workbook_error("The workbook does not contain any dues payer rows.")
    return rows
