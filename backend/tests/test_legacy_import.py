from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import Workbook
from sqlalchemy import func, select

from app.legacy_import import (
    ARTICLE_HEADERS,
    CANONICAL_DUPLICATE_ARTICLE_SHEETS,
    SPECIAL_SHEETS,
    SourceRow,
    _detected_content_type,
    _download_url,
    _attach_media_to_special_entry,
    _activity_certification_metadata,
    import_articles_and_specials,
    import_attachments,
    import_comments,
    index_local_attachment_files,
    load_article_workbook,
    normalize_author,
    redact_text,
    target_slug,
    validate_database_target,
)
from app.models.audit import LegacyImportRecord
from app.models.board import Board
from app.models.comment import Comment
from app.models.media import MediaAsset, PostAttachment
from app.models.post import Post
from scripts.verify_legacy_media import verify_legacy_media


def test_author_normalization_removes_duplicate_cohort_prefixes() -> None:
    assert normalize_author("[70기 70기_이혜진]") == ("이혜진", "70기")
    assert normalize_author("[서강대 72기 한다현]") == ("한다현", "72기")


def test_legacy_student_ids_are_redacted() -> None:
    value, findings = redact_text("1. A66071 정승연 데이터사이언스")
    assert value == "1. [학번 비공개] 정승연 데이터사이언스"
    assert findings == ["student_id"]


def test_free_board_sheet_is_only_a_staging_bucket() -> None:
    row = SourceRow(
        source_file="board_articles_ver2.xlsx",
        sheet="자유게시판",
        row_number=1,
        data={"writeId": "1", "boardName": "전공 커뮤니티"},
    )
    assert target_slug(row) == "community-major"
    assert "free-board" not in {target_slug(row)}


def test_staging_sheet_uses_target_board_name_as_visible_category(api) -> None:
    row = SourceRow(
        source_file="board_articles_ver2.xlsx",
        sheet="자유게시판",
        row_number=1,
        data={
            "writeId": "staging-1",
            "boardName": "전공 커뮤니티",
            "title": "전공 질문",
            "content": "본문",
            "writeUser": "72기 작성자",
            "cohort": "72기",
            "date": datetime(2025, 1, 1, 9, 0, 0),
        },
    )
    with api.session() as db:
        board = Board(
            name="전공 커뮤니티",
            slug="community-major",
            category="community",
            board_type="post",
            read_permission="user",
            write_permission="user",
        )
        db.add(board)
        db.commit()

        posts, _, _ = import_articles_and_specials(db, [row], [], apply=True)
        db.commit()

        assert posts["staging-1"].category == "전공 커뮤니티"


def test_import_target_rejects_source_database() -> None:
    with pytest.raises(RuntimeError, match="protected source database"):
        validate_database_target("postgresql+psycopg://postgres:postgres@db:5432/sogang_app")
    assert (
        validate_database_target(
            "postgresql+psycopg://postgres:postgres@db:5432/sogang_app_migration_review_20260802"
        )
        == "sogang_app_migration_review_20260802"
    )


def test_legacy_download_url_quotes_unicode_and_spaces() -> None:
    assert _download_url("https://example.com/첨부 파일.pdf?download=1") == (
        "https://example.com/%EC%B2%A8%EB%B6%80%20%ED%8C%8C%EC%9D%BC.pdf?download=1"
    )


def test_attachment_type_uses_file_signature_over_legacy_filename(tmp_path: Path) -> None:
    path = tmp_path / "mobile_upload_img.png"
    path.write_bytes(b"\xff\xd8\xff\xe0" + b"legacy-jpeg")

    assert _detected_content_type(path, "image/png", "image/png") == "image/jpeg"


def test_special_entry_keeps_banner_and_all_attachments() -> None:
    board = Board(
        name="Past councils",
        slug="gsa-past-councils",
        category="student-council",
        board_type="content",
        read_permission="user",
        write_permission="admin",
        metadata_json={
            "past_councils": [
                {"legacy_write_id": "4313487", "banner_image_url": "", "intro": "24대"}
            ]
        },
    )

    assert _attach_media_to_special_entry(
        board,
        collection_key="past_councils",
        article_id="4313487",
        media_url="/api/media/1",
    )
    assert _attach_media_to_special_entry(
        board,
        collection_key="past_councils",
        article_id="4313487",
        media_url="/api/media/2",
    )
    entry = board.metadata_json["past_councils"][0]
    assert entry["banner_image_url"] == "/api/media/1"
    assert entry["attachment_urls"] == ["/api/media/1", "/api/media/2"]


def test_activity_certification_extracts_historical_date_and_participants() -> None:
    metadata = _activity_certification_metadata(
        """[동아리명]
알바트로스냅

[이름 / 기수]
72기 김민경
70기_김형진
총 2명

[활동 날짜]
6월 6일 - 창경궁

[이미지 업로드]""",
        datetime(2026, 6, 27, 1, 41, 1),
        "알바트로스냅",
    )

    assert metadata["activity_date"] == "2026-06-06"
    assert metadata["participants"] == "72기 김민경, 70기 김형진"
    assert metadata["legacy_activity_name"] == "알바트로스냅"


def test_study_activity_extracts_two_digit_year_and_ignores_major_suffix() -> None:
    metadata = _activity_certification_metadata(
        """[스터디원 이름/기수/학과]
김가현 74기 인공지능
허명진 / 68기 / 데이터사이언스

[스터디 날짜 및 시간 ]
26.05.30

[스터디 내용]
기말고사 준비""",
        datetime(2026, 6, 1, 9, 0, 0),
        "딥러닝 스터디",
    )

    assert metadata["activity_date"] == "2026-05-30"
    assert metadata["participants"] == "74기 김가현, 68기 허명진"


def test_activity_participants_skip_template_notes_and_remove_student_ids() -> None:
    metadata = _activity_certification_metadata(
        """[이름 / 기수]
[ 7/18 모임참석]
69기_김준호
64 강현구

[첫 활동일]
7/18""",
        datetime(2025, 7, 19, 9, 0, 0),
        "서뽈링",
    )
    assert metadata["participants"] == "69기 김준호, 64기 강현구"

    study = _activity_certification_metadata(
        """[스터디원 이름/기수/학과]
1. A66071 정승연 데이터사이언스
2. A65002 강지은 데이터사이언스

[스터디 날짜 및 시간 ]
2023.03.14 20:00""",
        datetime(2023, 3, 20, 9, 0, 0),
        "AI 스터디",
    )
    assert study["participants"] == "66기 정승연, 65기 강지은"
    assert "A66071" not in study["participants"]


def test_legacy_activity_extracts_grouped_cohorts_and_inline_date() -> None:
    metadata = _activity_certification_metadata(
        """[이름 / 기수]
66기-최철, 정승연
67기-권영환, 선용준

모임일자: 2024.03.15""",
        datetime(2024, 3, 20, 9, 0, 0),
        "동아리 활동",
    )

    assert metadata["activity_date"] == "2024-03-15"
    assert metadata["participants"] == "66기 최철, 66기 정승연, 67기 권영환, 67기 선용준"


def test_legacy_activity_recovers_cohort_from_student_id_without_storing_id() -> None:
    metadata = _activity_certification_metadata(
        """4월 15일 모임에 참석했습니다.
주정헌(A701234)
A725678 변유철""",
        datetime(2022, 4, 20, 9, 0, 0),
        "스터디 활동",
    )

    assert metadata["activity_date"] == "2022-04-15"
    assert metadata["participants"] == "70기 주정헌, 72기 변유철"
    assert "A701234" not in metadata["participants"]


def test_headerless_staging_sheet_and_duplicate_article_are_parsed(tmp_path: Path) -> None:
    path = tmp_path / "board_articles_ver2.xlsx"
    workbook = Workbook()
    staging = workbook.active
    staging.title = "자유게시판"
    first = [None] * len(ARTICLE_HEADERS)
    first[ARTICLE_HEADERS.index("boardName")] = "전공 커뮤니티"
    first[ARTICLE_HEADERS.index("writeId")] = 101
    first[ARTICLE_HEADERS.index("title")] = "첫 글"
    staging.append(first)
    notices = workbook.create_sheet("기타공지")
    notices.append(ARTICLE_HEADERS)
    second = [None] * len(ARTICLE_HEADERS)
    second[ARTICLE_HEADERS.index("boardName")] = "전체공지"
    second[ARTICLE_HEADERS.index("writeId")] = 101
    second[ARTICLE_HEADERS.index("title")] = "중복 글"
    notices.append(second)
    attachments = workbook.create_sheet("첨부파일")
    attachments.append(
        [
            "writeId", "boardId", "attach_type", "sequence", "subject", "fileStorageId",
            "contentType", "attach_id", "link_url", "regiDatetime", "fileSize",
        ]
    )
    attachments.append([101, "board", "image", 0, "photo.png", 9001, "BOARD_ATC_IMAGE", 9001, None, None, 3])
    workbook.save(path)

    articles, parsed_attachments, duplicates = load_article_workbook(path)
    assert [row.source_id for row in articles] == ["101"]
    assert articles[0].sheet == "자유게시판"
    assert len(parsed_attachments) == 1
    assert len(duplicates) == 1


def test_known_duplicate_article_prefers_canonical_photo_album_sheet(tmp_path: Path) -> None:
    source_id, canonical_sheet = next(iter(CANONICAL_DUPLICATE_ARTICLE_SHEETS.items()))
    path = tmp_path / "board_articles_ver2.xlsx"
    workbook = Workbook()
    noncanonical = workbook.active
    noncanonical.title = "기타공지"
    noncanonical.append(ARTICLE_HEADERS)
    first = [None] * len(ARTICLE_HEADERS)
    first[ARTICLE_HEADERS.index("writeId")] = source_id
    first[ARTICLE_HEADERS.index("title")] = "duplicate"
    noncanonical.append(first)
    canonical = workbook.create_sheet(canonical_sheet)
    canonical.append(ARTICLE_HEADERS)
    second = [None] * len(ARTICLE_HEADERS)
    second[ARTICLE_HEADERS.index("writeId")] = source_id
    second[ARTICLE_HEADERS.index("title")] = "canonical album"
    canonical.append(second)
    workbook.save(path)

    articles, _, duplicates = load_article_workbook(path)

    assert len(articles) == 1
    assert articles[0].sheet == canonical_sheet
    assert articles[0].data["title"] == "canonical album"
    assert len(duplicates) == 1
    assert duplicates[0].sheet == "기타공지"


def test_local_attachment_import_links_one_copy_per_post_and_serves_it(
    api,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    from app.config import settings

    source_dir = tmp_path / "attachment-source"
    public_dir = tmp_path / "public-media"
    private_dir = tmp_path / "private-media"
    source_dir.mkdir()
    jpeg = b"\xff\xd8\xff\xe0" + b"legacy-jpeg-content"
    (source_dir / "9001.jpg").write_bytes(jpeg)
    (source_dir / "9002.jpg").write_bytes(jpeg)
    monkeypatch.setattr(settings, "media_upload_dir", public_dir)
    monkeypatch.setattr(settings, "media_private_upload_dir", private_dir)

    rows = [
        SourceRow(
            "board_articles_ver2.xlsx",
            "첨부파일",
            row_number,
            {
                "writeId": "article-1",
                "fileStorageId": storage_id,
                "subject": filename,
                "sequence": sequence,
                "regiDatetime": datetime(2025, 1, 1, 9, 0, 0),
            },
        )
        for row_number, storage_id, filename, sequence in (
            (2, "9001", "first.jpg", 0),
            (3, "9002", "duplicate.jpg", 1),
        )
    ]

    assert set(index_local_attachment_files(source_dir)) == {"9001", "9002"}
    with api.session() as db:
        post = db.get(Post, 3)
        stats = import_attachments(
            db,
            rows,
            {},
            {"article-1": post},
            public_media_dir=public_dir,
            private_media_dir=private_dir,
            attachment_source_dir=source_dir,
            apply=True,
            skip_downloads=False,
        )
        db.commit()

        imported_media = db.scalars(
            select(MediaAsset)
            .where(MediaAsset.stored_filename.like("legacy-900%"))
            .order_by(MediaAsset.id)
        ).all()
        imported_links = db.scalars(
            select(PostAttachment)
            .where(
                PostAttachment.post_id == post.id,
                PostAttachment.media_id.in_([media.id for media in imported_media]),
            )
        ).all()
        verification = verify_legacy_media(
            db,
            public_media_dir=public_dir,
            private_media_dir=private_dir,
        )

    assert len(imported_media) == 2
    assert len(imported_links) == 1
    assert stats["deduplicated_post_attachment_links"] == 1
    assert all(media.url == f"/api/media/{media.id}/access-url" for media in imported_media)
    assert all((public_dir / media.stored_filename).is_file() for media in imported_media)
    assert verification["status"] == "ok"
    assert verification["counts"]["verified_records"] == 2
    assert verification["counts"]["verified_files"] == 2

    detail = api.client.get("/api/posts/3", headers=api.headers["owner"])
    assert detail.status_code == 200
    imported_payloads = [
        attachment
        for attachment in detail.json()["data"]["attachments"]
        if attachment["id"] in {media.id for media in imported_media}
    ]
    assert len(imported_payloads) == 1

    access = api.client.get(
        imported_payloads[0]["url"],
        headers=api.headers["owner"],
    )
    assert access.status_code == 200
    file_response = api.client.get(access.json()["data"]["url"])
    assert file_response.status_code == 200
    assert file_response.content == jpeg


def test_local_faq_image_is_returned_and_authorized(
    api,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    from app.config import settings

    source_dir = tmp_path / "faq-source"
    public_dir = tmp_path / "faq-public"
    private_dir = tmp_path / "faq-private"
    source_dir.mkdir()
    png = b"\x89PNG\r\n\x1a\n" + b"legacy-faq-image"
    (source_dir / "4322320.png").write_bytes(png)
    monkeypatch.setattr(settings, "media_upload_dir", public_dir)
    monkeypatch.setattr(settings, "media_private_upload_dir", private_dir)
    faq_sheet = next(sheet for sheet, kind in SPECIAL_SHEETS.items() if kind == "faq")
    article = SourceRow(
        "board_articles_ver2.xlsx",
        faq_sheet,
        2,
        {"writeId": "faq-1", "title": "FAQ question", "content": "FAQ answer"},
    )
    attachment = SourceRow(
        "board_articles_ver2.xlsx",
        "첨부파일",
        3,
        {
            "writeId": "faq-1",
            "fileStorageId": "4322320",
            "subject": "faq.png",
            "sequence": 0,
        },
    )

    with api.session() as db:
        posts, _, _ = import_articles_and_specials(db, [article], [], apply=True)
        stats = import_attachments(
            db,
            [attachment],
            {},
            posts,
            public_media_dir=public_dir,
            private_media_dir=private_dir,
            attachment_source_dir=source_dir,
            apply=True,
            skip_downloads=False,
        )
        db.commit()

    assert stats["created_attachments"] == 1
    response = api.client.get("/api/faqs", headers=api.headers["owner"])
    assert response.status_code == 200
    faq_payload = next(item for item in response.json()["data"] if item["question"] == "FAQ question")
    assert len(faq_payload["attachments"]) == 1
    media = faq_payload["attachments"][0]
    assert media["content_type"] == "image/png"

    access = api.client.get(media["url"], headers=api.headers["owner"])
    assert access.status_code == 200
    file_response = api.client.get(access.json()["data"]["url"])
    assert file_response.status_code == 200
    assert file_response.content == png


def test_unsupported_legacy_download_is_archived_without_exposing_media(api, tmp_path: Path) -> None:
    source_dir = tmp_path / "unsupported-source"
    public_dir = tmp_path / "unsupported-public"
    private_dir = tmp_path / "unsupported-private"
    source_dir.mkdir()
    (source_dir / "12621604.zip").write_bytes(b"PK\x03\x04legacy-archive")
    row = SourceRow(
        "board_articles_ver2.xlsx",
        "첨부파일",
        2,
        {
            "writeId": "article-1",
            "fileStorageId": "12621604",
            "subject": "source.zip",
            "sequence": 0,
        },
    )

    with api.session() as db:
        post = db.get(Post, 3)
        media_count_before = db.scalar(select(func.count(MediaAsset.id)))
        stats = import_attachments(
            db,
            [row],
            {},
            {"article-1": post},
            public_media_dir=public_dir,
            private_media_dir=private_dir,
            attachment_source_dir=source_dir,
            apply=True,
            skip_downloads=False,
        )
        db.commit()
        media_count_after = db.scalar(select(func.count(MediaAsset.id)))
        record = db.scalar(
            select(LegacyImportRecord).where(
                LegacyImportRecord.entity_type == "attachment",
                LegacyImportRecord.source_id == "12621604",
            )
        )

    assert stats["archived_unsupported_attachments"] == 1
    assert media_count_after == media_count_before
    assert record.status == "archived"
    assert record.reason == "legacy_attachment_type_not_supported"


def test_article_import_is_idempotent(api) -> None:
    with api.session() as db:
        db.add(
            Board(
                name="Academic notices",
                slug="academic-notices",
                category="notices",
                board_type="notice",
                read_permission="user",
                write_permission="admin",
            )
        )
        db.commit()
        row = SourceRow(
            source_file="board_articles_ver2.xlsx",
            sheet="학사공지",
            row_number=2,
            data={
                "writeId": "4313612",
                "boardName": "전체공지",
                "title": "이관 테스트",
                "content": "동일한 원본을 다시 실행해도 한 건만 남아야 합니다.",
                "writeUser": "72기 한다현",
                "cohort": "72기",
                "date": datetime(2025, 1, 1, 9, 0, 0),
                "updateDate": datetime(2025, 1, 1, 9, 0, 0),
            },
        )

        imported_posts, first_stats, _ = import_articles_and_specials(db, [row], [], apply=True, limit=1)
        db.commit()
        imported_post = imported_posts["4313612"]
        expected_snapshot = (
            imported_post.metadata_json["legacy_author"],
            imported_post.metadata_json["legacy_author_cohort"],
        )
        imported_post.author_nickname_snapshot = None
        imported_post.author_cohort_snapshot = None
        db.commit()
        _, second_stats, _ = import_articles_and_specials(db, [row], [], apply=True, limit=1)
        db.commit()
        _, third_stats, _ = import_articles_and_specials(db, [row], [], apply=True, limit=1)
        db.commit()

        assert first_stats["created_posts"] == 1
        assert second_stats["updated_posts"] == 1
        assert third_stats["unchanged_posts"] == 1
        db.refresh(imported_post)
        assert (
            imported_post.author_nickname_snapshot,
            imported_post.author_cohort_snapshot,
        ) == expected_snapshot
        assert db.scalar(select(func.count(Post.id)).where(Post.title == "이관 테스트")) == 1
        assert db.scalar(select(func.count(LegacyImportRecord.id))) == 1


def test_special_metadata_rerun_preserves_imported_attachment_references(api) -> None:
    special_sheet = next(sheet for sheet, kind in SPECIAL_SHEETS.items() if kind == "cohort_leader")
    row = SourceRow(
        source_file="board_articles_ver2.xlsx",
        sheet=special_sheet,
        row_number=2,
        data={
            "writeId": "9001",
            "title": "72기 기장단",
            "content": "72기 기장 홍길동\n반갑습니다.",
            "writeUser": "72기 홍길동",
            "cohort": "72기",
        },
    )
    with api.session() as db:
        db.add(
            Board(
                name="Cohort leaders",
                slug="gsa-cohort-leaders",
                category="gsa",
                board_type="organization_intro",
                read_permission="user",
                write_permission="admin",
            )
        )
        db.commit()
        import_articles_and_specials(db, [row], [], apply=True)
        db.commit()
        board = db.scalar(select(Board).where(Board.slug == "gsa-cohort-leaders"))
        metadata = dict(board.metadata_json)
        entries = list(metadata["cohort_leaders"])
        entries[0] = {
            **entries[0],
            "banner_image_url": "/api/media/11",
            "attachment_urls": ["/api/media/11", "/api/media/12"],
        }
        metadata["cohort_leaders"] = entries
        board.metadata_json = metadata
        db.commit()

        import_articles_and_specials(db, [row], [], apply=True)
        db.commit()
        db.refresh(board)
        entry = board.metadata_json["cohort_leaders"][0]
        assert entry["banner_image_url"] == "/api/media/11"
        assert entry["attachment_urls"] == ["/api/media/11", "/api/media/12"]


def test_comment_parent_mapping_archives_depth_over_two(api) -> None:
    with api.session() as db:
        board = Board(
            name="Major community",
            slug="community-major",
            category="community",
            board_type="post",
            read_permission="user",
            write_permission="user",
        )
        db.add(board)
        db.flush()
        post = Post(
            board_id=board.id,
            author_id=1,
            title="Legacy post",
            content="Body",
            metadata_json={"legacy_write_id": "article-1"},
        )
        db.add(post)
        db.flush()
        rows = [
            SourceRow(
                "comments.xlsx",
                "comments",
                index + 2,
                {
                    "writeId": source_id,
                    "article_id": "article-1",
                    "parent_comment_id": parent_id,
                    "content": content,
                    "date": datetime(2025, 1, index + 1, 9, 0, 0),
                    "write_user": "72기 작성자",
                    "cohort": "72기",
                },
            )
            for index, (source_id, parent_id, content) in enumerate(
                [("c1", "", "root"), ("c2", "c1", "reply"), ("c3", "c2", "too deep")]
            )
        ]

        stats = import_comments(db, rows, {"article-1": post}, apply=True, reconcile_untracked=False)
        db.commit()

        imported_comments = db.scalars(
            select(Comment).where(Comment.post_id == post.id).order_by(Comment.id)
        ).all()
        for comment in imported_comments:
            comment.author_nickname_snapshot = None
            comment.author_cohort_snapshot = None
        db.commit()
        rerun_stats = import_comments(db, rows, {"article-1": post}, apply=True, reconcile_untracked=False)
        db.commit()

        assert stats["created_comments"] == 2
        assert rerun_stats["created_comments"] == 0
        assert stats["over_depth_reply_comments"] == 1
        assert db.scalar(select(func.count(Comment.id)).where(Comment.post_id == post.id)) == 2
        for comment in imported_comments:
            db.refresh(comment)
            assert comment.author_nickname_snapshot is not None
            assert comment.author_cohort_snapshot is not None
        archived = db.scalar(
            select(LegacyImportRecord).where(
                LegacyImportRecord.entity_type == "comment",
                LegacyImportRecord.source_id == "c3",
            )
        )
        assert archived.status == "archived"
        assert archived.reason == "comment_depth_exceeds_two"


def test_legacy_import_review_api_is_admin_only(api) -> None:
    with api.session() as db:
        db.add(
            LegacyImportRecord(
                source_file="board_articles_ver2.xlsx",
                source_sheet="자유게시판",
                source_row=1,
                entity_type="article",
                source_id="101",
                source_hash="a" * 64,
                action="review",
                status="unmapped",
                reason="no_approved_board_mapping",
                redacted_details={"title": "검토 글"},
            )
        )
        db.commit()

    assert api.client.get("/api/admin/legacy-import/summary", headers=api.headers["owner"]).status_code == 403
    response = api.client.get("/api/admin/legacy-import/summary", headers=api.headers["admin"])
    assert response.status_code == 200
    assert response.json()["data"] == [
        {"entity_type": "article", "status": "unmapped", "action": "review", "count": 1}
    ]
