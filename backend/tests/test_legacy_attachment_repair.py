from __future__ import annotations

import importlib
import importlib.util
import io
import json
from datetime import datetime
from pathlib import Path
import zipfile

import pytest
from openpyxl import Workbook, load_workbook
from sqlalchemy import select

from app.models.audit import LegacyImportRecord
from app.models.media import MediaAsset, PostAttachment
from app.models.post import Post


def _repair_module():
    spec = importlib.util.find_spec("app.legacy_attachment_repair")
    assert spec is not None, "legacy attachment repair service must exist"
    return importlib.import_module("app.legacy_attachment_repair")


def _repair_cli_module():
    script = Path(__file__).resolve().parents[1] / "scripts" / "repair_legacy_attachments.py"
    spec = importlib.util.spec_from_file_location("repair_legacy_attachments_cli", script)
    assert spec is not None and spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def _zip_bytes() -> bytes:
    output = io.BytesIO()
    with zipfile.ZipFile(output, "w") as archive:
        archive.writestr("answer.txt", "exam answer")
    return output.getvalue()


def _source_bytes(extension: str) -> bytes:
    if extension in {".hwp", ".doc"}:
        return b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1" + (b"\x00" * 128) + b"HWP Document File"
    if extension == ".zip":
        return _zip_bytes()
    if extension == ".txt":
        return "기말고사 예상문제\n".encode()
    if extension == ".ipynb":
        return json.dumps({"cells": [], "metadata": {}, "nbformat": 4, "nbformat_minor": 5}).encode()
    if extension == ".mp4":
        return b"\x00\x00\x00\x18ftypmp42legacy-video"
    raise AssertionError(f"Unsupported fixture extension: {extension}")


def _write_attachment_workbook(path: Path, rows: list[tuple[str, str, str]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "첨부파일"
    sheet.append(
        [
            "writeId",
            "boardId",
            "attach_type",
            "sequence",
            "subject",
            "fileStorageId",
            "contentType",
            "attach_id",
            "link_url",
            "regiDatetime",
            "fileSize",
        ]
    )
    for sequence, (article_id, storage_id, extension) in enumerate(rows):
        sheet.append(
            [
                article_id,
                None,
                "file",
                sequence,
                f"exam-material{extension}",
                storage_id,
                None,
                None,
                None,
                None,
                None,
            ]
        )
    workbook.save(path)


def _prepare_source(tmp_path: Path, rows: list[tuple[str, str, str]]) -> tuple[Path, Path, Path, Path]:
    articles = tmp_path / "board_articles_ver3.xlsx"
    source = tmp_path / "attachments"
    public = tmp_path / "public"
    private = tmp_path / "private"
    source.mkdir()
    public.mkdir()
    private.mkdir()
    _write_attachment_workbook(articles, rows)
    for _, storage_id, extension in rows:
        body = _source_bytes(extension)
        if extension == ".txt":
            body += storage_id.encode()
        (source / f"{storage_id}{extension}").write_bytes(body)
    return articles, source, public, private


def test_selected_repair_preserves_existing_hwp_and_inserts_only_missing_attachments(
    api,
    tmp_path: Path,
) -> None:
    repair = _repair_module()
    expected_articles = repair.QA_175_176_EXPECTED_ARTICLE_IDS
    extensions = {
        "10946091": ".hwp",
        "12621604": ".zip",
        "12621345": ".zip",
        "12358989": ".zip",
        "10946120": ".zip",
        "12359030": ".txt",
        "12359031": ".txt",
        "12358946": ".ipynb",
    }
    stored_extensions = {**extensions, "10946091": ".doc"}
    rows = [
        (article_id, storage_id, extensions[storage_id])
        for storage_id, article_id in expected_articles.items()
    ]
    articles, source, public, private = _prepare_source(tmp_path, rows)
    old_hwp = public / "legacy-10946091.doc"
    old_hwp.write_bytes(_source_bytes(".hwp"))

    with api.session() as db:
        unique_article_ids = tuple(dict.fromkeys(expected_articles.values()))
        posts_by_article_id = {}
        template_post = db.get(Post, 3)
        for index, article_id in enumerate(unique_article_ids):
            post = template_post if index == 0 else Post(
                board_id=template_post.board_id,
                author_id=template_post.author_id,
                author_nickname_snapshot=template_post.author_nickname_snapshot,
                author_cohort_snapshot=template_post.author_cohort_snapshot,
                title=f"Legacy repair parent {article_id}",
                content="Legacy repair fixture",
            )
            post.status = "published"
            post.deleted_at = None
            post.metadata_json = {"legacy_write_id": article_id}
            db.add(post)
            db.flush()
            posts_by_article_id[article_id] = post
        hwp_post = posts_by_article_id[expected_articles["10946091"]]
        media = MediaAsset(
            owner_id=hwp_post.author_id,
            original_filename="legacy-10946091",
            stored_filename=old_hwp.name,
            content_type="application/msword",
            file_size=old_hwp.stat().st_size,
            url=None,
            is_private=False,
            status="ready",
        )
        db.add(media)
        db.flush()
        media.url = f"/api/media/{media.id}/access-url"
        db.add(PostAttachment(post_id=hwp_post.id, media_id=media.id, sort_order=0))
        legacy_record = LegacyImportRecord(
            source_file=articles.name,
            source_sheet="첨부파일",
            source_row=3,
            entity_type="attachment",
            source_id="12621604",
            source_parent_id=expected_articles["12621604"],
            source_hash="a" * 64,
            action="archived",
            status="archived",
            target_table=None,
            target_id=None,
            reason="preserve-existing-ledger",
            redacted_details={"preserve": True},
        )
        db.add(legacy_record)
        db.commit()
        original_media_id = media.id
        original_media_values = (
            media.original_filename,
            media.stored_filename,
            media.content_type,
            media.file_size,
            media.url,
            media.is_private,
            media.status,
        )
        original_media_count = len(db.scalars(select(MediaAsset)).all())
        original_attachment_count = len(db.scalars(select(PostAttachment)).all())
        original_legacy_record_values = (
            legacy_record.source_row,
            legacy_record.source_parent_id,
            legacy_record.source_hash,
            legacy_record.action,
            legacy_record.status,
            legacy_record.target_table,
            legacy_record.target_id,
            legacy_record.reason,
            legacy_record.redacted_details,
            legacy_record.updated_at,
        )

        result = repair.repair_legacy_attachments(
            db,
            articles_xlsx=articles,
            attachment_source_dir=source,
            public_media_dir=public,
            private_media_dir=private,
            storage_ids=tuple(expected_articles),
            expected_article_ids=expected_articles,
            apply=True,
        )

        selected_media = [
            item
            for item in db.scalars(select(MediaAsset)).all()
            if item.stored_filename.startswith("legacy-")
            and item.stored_filename.split(".", 1)[0].removeprefix("legacy-") in expected_articles
        ]
        db.refresh(media)
        db.refresh(legacy_record)
        final_media_count = len(db.scalars(select(MediaAsset)).all())
        final_attachment_count = len(db.scalars(select(PostAttachment)).all())
        final_legacy_record_values = (
            legacy_record.source_row,
            legacy_record.source_parent_id,
            legacy_record.source_hash,
            legacy_record.action,
            legacy_record.status,
            legacy_record.target_table,
            legacy_record.target_id,
            legacy_record.reason,
            legacy_record.redacted_details,
            legacy_record.updated_at,
        )

    assert result["mode"] == "apply"
    assert result["selected_storage_ids"] == list(expected_articles)
    assert {
        item["storage_id"]: item["legacy_article_id"]
        for item in result["planned_files"]
    } == expected_articles
    assert len({item["post_id"] for item in result["planned_files"]}) == 7
    assert all(item["post_status"] == "published" for item in result["planned_files"])
    assert result["attachment_stats"] == {"created_attachments": 7}
    assert result["validated_existing_storage_ids"] == ["10946091"]
    assert old_hwp.exists()
    assert old_hwp.read_bytes() == _source_bytes(".hwp")
    assert media.id == original_media_id
    assert (
        media.original_filename,
        media.stored_filename,
        media.content_type,
        media.file_size,
        media.url,
        media.is_private,
        media.status,
    ) == original_media_values
    assert final_media_count == original_media_count + 7
    assert final_attachment_count == original_attachment_count + 7
    assert final_legacy_record_values == original_legacy_record_values
    assert len(selected_media) == 8
    assert {
        item.stored_filename.removeprefix("legacy-").split(".", 1)[0]: Path(item.stored_filename).suffix
        for item in selected_media
    } == stored_extensions


def test_selected_repair_rejects_preexisting_insert_target_without_updates(
    api,
    tmp_path: Path,
) -> None:
    repair = _repair_module()
    rows = [("article-1", "12621604", ".zip")]
    articles, source, public, private = _prepare_source(tmp_path, rows)
    existing_path = public / "legacy-12621604.zip"
    existing_bytes = _source_bytes(".zip")
    existing_path.write_bytes(existing_bytes)

    with api.session() as db:
        post = db.get(Post, 3)
        post.status = "published"
        post.deleted_at = None
        post.metadata_json = {"legacy_write_id": "article-1"}
        media = MediaAsset(
            owner_id=post.author_id,
            original_filename="do-not-update.zip",
            stored_filename=existing_path.name,
            content_type="application/zip",
            file_size=existing_path.stat().st_size,
            url=None,
            is_private=False,
            status="ready",
        )
        db.add(media)
        db.flush()
        media.url = f"/api/media/{media.id}/access-url"
        db.add(PostAttachment(post_id=post.id, media_id=media.id, sort_order=9))
        db.commit()
        media_id = media.id

        with pytest.raises(ValueError, match="insert-only repair targets already exist: 12621604"):
            repair.repair_legacy_attachments(
                db,
                articles_xlsx=articles,
                attachment_source_dir=source,
                public_media_dir=public,
                private_media_dir=private,
                storage_ids=("12621604",),
                apply=True,
            )

        unchanged = db.get(MediaAsset, media_id)
        assert unchanged.original_filename == "do-not-update.zip"
        assert unchanged.stored_filename == "legacy-12621604.zip"
        assert unchanged.content_type == "application/zip"
        link = db.scalar(
            select(PostAttachment).where(
                PostAttachment.post_id == post.id,
                PostAttachment.media_id == media_id,
            )
        )
        assert link.sort_order == 9

    assert existing_path.read_bytes() == existing_bytes


@pytest.mark.parametrize("invalid_bytes", [b"", b"%PDF-1.7\nnot-a-zip", b"\x00\xff\x00\xff" * 64])
def test_dry_run_rejects_invalid_selected_file_without_writes(
    api,
    tmp_path: Path,
    invalid_bytes: bytes,
) -> None:
    repair = _repair_module()
    rows = [
        ("article-1", "broken-zip", ".zip"),
    ]
    articles, source, public, private = _prepare_source(tmp_path, rows)
    (source / "broken-zip.zip").write_bytes(invalid_bytes)

    with api.session() as db:
        post = db.get(Post, 3)
        post.metadata_json = {"legacy_write_id": "article-1"}
        db.commit()

        with pytest.raises(ValueError, match="unsupported or invalid local files: broken-zip"):
            repair.repair_legacy_attachments(
                db,
                articles_xlsx=articles,
                attachment_source_dir=source,
                public_media_dir=public,
                private_media_dir=private,
                storage_ids=("broken-zip",),
                apply=False,
            )

    assert not tuple(public.iterdir())
    assert not tuple(private.iterdir())


def test_dry_run_rejects_selected_file_above_apply_size_limit(api, tmp_path: Path) -> None:
    repair = _repair_module()
    rows = [("article-1", "12621604", ".zip")]
    articles, source, public, private = _prepare_source(tmp_path, rows)

    with api.session() as db:
        post = db.get(Post, 3)
        post.metadata_json = {"legacy_write_id": "article-1"}
        db.commit()

        with pytest.raises(ValueError, match="local files exceed 32 bytes: 12621604"):
            repair.repair_legacy_attachments(
                db,
                articles_xlsx=articles,
                attachment_source_dir=source,
                public_media_dir=public,
                private_media_dir=private,
                storage_ids=("12621604",),
                maximum_bytes=32,
                apply=False,
            )


def test_dry_run_rejects_deleted_legacy_parent_post(api, tmp_path: Path) -> None:
    repair = _repair_module()
    rows = [("article-1", "12621604", ".zip")]
    articles, source, public, private = _prepare_source(tmp_path, rows)

    with api.session() as db:
        post = db.get(Post, 3)
        post.status = "published"
        post.deleted_at = datetime.utcnow()
        post.metadata_json = {"legacy_write_id": "article-1"}
        db.commit()

        with pytest.raises(ValueError, match="Published legacy parent posts were not found: article-1"):
            repair.repair_legacy_attachments(
                db,
                articles_xlsx=articles,
                attachment_source_dir=source,
                public_media_dir=public,
                private_media_dir=private,
                storage_ids=("12621604",),
                apply=False,
            )


def test_dry_run_rejects_duplicate_published_legacy_parent_posts(api, tmp_path: Path) -> None:
    repair = _repair_module()
    rows = [("article-1", "12621604", ".zip")]
    articles, source, public, private = _prepare_source(tmp_path, rows)

    with api.session() as db:
        first = db.get(Post, 3)
        second = db.get(Post, 4)
        first.status = "published"
        first.deleted_at = None
        second.status = "published"
        second.deleted_at = None
        first.metadata_json = {"legacy_write_id": "article-1"}
        second.metadata_json = {"legacy_write_id": "article-1"}
        db.commit()

        with pytest.raises(ValueError, match="Multiple published legacy parent posts: article-1"):
            repair.repair_legacy_attachments(
                db,
                articles_xlsx=articles,
                attachment_source_dir=source,
                public_media_dir=public,
                private_media_dir=private,
                storage_ids=("12621604",),
                apply=False,
            )


def test_dry_run_rejects_explicit_video_selection(api, tmp_path: Path) -> None:
    repair = _repair_module()
    rows = [("article-1", "video-1", ".mp4")]
    articles, source, public, private = _prepare_source(tmp_path, rows)

    with api.session() as db:
        post = db.get(Post, 3)
        post.metadata_json = {"legacy_write_id": "article-1"}
        db.commit()

        with pytest.raises(ValueError, match="unsupported repair extensions: video-1"):
            repair.repair_legacy_attachments(
                db,
                articles_xlsx=articles,
                attachment_source_dir=source,
                public_media_dir=public,
                private_media_dir=private,
                storage_ids=("video-1",),
                apply=False,
            )


def test_target_descriptor_hashes_database_query_secrets(monkeypatch) -> None:
    cli = _repair_cli_module()
    monkeypatch.setattr(
        cli.settings,
        "database_url",
        "postgresql+psycopg://operator:password@db/aisw?sslpassword=do-not-log&options=-csearch_path%3Dquery-secret-schema",
    )
    monkeypatch.setattr(cli.settings, "public_api_url", "https://api.example")

    descriptor = cli._target_descriptor()
    serialized = json.dumps(descriptor, sort_keys=True)

    assert "do-not-log" not in serialized
    assert "query-secret-schema" not in serialized
    assert descriptor["database_query_sha256"]
    assert "database_query" not in descriptor


def test_apply_target_guard_binds_confirmation_to_production_endpoint() -> None:
    cli = _repair_cli_module()
    target = {
        "app_environment": "production",
        "database_host": "db",
        "database_port": 5432,
        "database_name": "aisw",
        "public_api_url": "https://34.50.35.119/api",
    }
    fingerprint = cli._target_fingerprint(target)

    assert cli._apply_target_errors(
        target,
        expected_database_name="aisw",
        expected_target_fingerprint=fingerprint,
    ) == []

    staging_target = {**target, "public_api_url": "https://staging.example/api"}
    assert cli._apply_target_errors(
        staging_target,
        expected_database_name="aisw",
        expected_target_fingerprint=fingerprint,
    ) == ["--expected-target-fingerprint must match the preceding dry-run target"]

    development_target = {**target, "app_environment": "development"}
    assert "APP_ENVIRONMENT must be production" in cli._apply_target_errors(
        development_target,
        expected_database_name="aisw",
        expected_target_fingerprint=cli._target_fingerprint(development_target),
    )

    query_target = {**target, "database_query_sha256": "a" * 64}
    assert cli._target_fingerprint(query_target) != fingerprint


def test_plan_fingerprint_binds_source_bytes_and_media_directories(tmp_path: Path) -> None:
    cli = _repair_cli_module()
    planned = {
        "storage_id": "10946091",
        "source_filename": "10946091.hwp",
        "source_size": 10,
        "source_sha256": "a" * 64,
        "detected_content_type": "application/x-hwp",
        "target_extension": ".hwp",
        "legacy_article_id": "6471114",
        "post_id": 414,
        "post_status": "published",
        "workbook_row_sha256": "d" * 64,
        "workbook_subject": "exam.hwp",
        "workbook_sequence": 0,
        "workbook_link_url_sha256": "e" * 64,
        "reference_url_sha256": "f" * 64,
        "expected_original_filename": "exam.hwp",
        "expected_is_private": False,
    }
    result = {"selected_storage_ids": ["10946091"], "planned_files": [planned]}
    public = tmp_path / "public"
    private = tmp_path / "private"
    fingerprint = cli._plan_fingerprint(
        result,
        target_fingerprint="b" * 64,
        public_media_dir=public,
        private_media_dir=private,
    )

    changed_source = {
        **result,
        "planned_files": [{**planned, "source_sha256": "c" * 64}],
    }
    assert cli._plan_fingerprint(
        changed_source,
        target_fingerprint="b" * 64,
        public_media_dir=public,
        private_media_dir=private,
    ) != fingerprint
    changed_privacy = {
        **result,
        "planned_files": [
            {
                **planned,
                "reference_url_sha256": "1" * 64,
                "expected_is_private": True,
            }
        ],
    }
    assert cli._plan_fingerprint(
        changed_privacy,
        target_fingerprint="b" * 64,
        public_media_dir=public,
        private_media_dir=private,
    ) != fingerprint
    assert cli._plan_fingerprint(
        result,
        target_fingerprint="b" * 64,
        public_media_dir=tmp_path / "wrong-public",
        private_media_dir=private,
    ) != fingerprint


def test_repair_preview_hashes_signed_source_urls(api, tmp_path: Path) -> None:
    repair = _repair_module()
    rows = [("article-1", "12359030", ".txt")]
    articles, source, public, private = _prepare_source(tmp_path, rows)
    secret = "signed-secret-token"
    workbook = load_workbook(articles)
    sheet = workbook["첨부파일"]
    sheet.cell(row=2, column=9, value=f"https://files.example/public/file.txt?X-Amz-Signature={secret}")
    workbook.save(articles)

    with api.session() as db:
        post = db.get(Post, 3)
        post.status = "published"
        post.deleted_at = None
        post.metadata_json = {"legacy_write_id": "article-1"}
        db.commit()
        result = repair.repair_legacy_attachments(
            db,
            articles_xlsx=articles,
            attachment_source_dir=source,
            public_media_dir=public,
            private_media_dir=private,
            storage_ids=("12359030",),
            apply=False,
        )

    serialized = json.dumps(result, sort_keys=True)
    assert secret not in serialized
    assert result["planned_files"][0]["workbook_link_url_sha256"]
    assert "workbook_link_url" not in result["planned_files"][0]


def test_apply_rejects_source_changed_after_confirmed_dry_run(api, tmp_path: Path) -> None:
    repair = _repair_module()
    rows = [("article-1", "12359030", ".txt")]
    articles, source, public, private = _prepare_source(tmp_path, rows)

    with api.session() as db:
        post = db.get(Post, 3)
        post.status = "published"
        post.deleted_at = None
        post.metadata_json = {"legacy_write_id": "article-1"}
        db.commit()
        preview = repair.repair_legacy_attachments(
            db,
            articles_xlsx=articles,
            attachment_source_dir=source,
            public_media_dir=public,
            private_media_dir=private,
            storage_ids=("12359030",),
            apply=False,
        )
        (source / "12359030.txt").write_text("changed after dry-run", encoding="utf-8")

        with pytest.raises(ValueError, match="plan changed after the confirmed dry-run"):
            repair.repair_legacy_attachments(
                db,
                articles_xlsx=articles,
                attachment_source_dir=source,
                public_media_dir=public,
                private_media_dir=private,
                storage_ids=("12359030",),
                expected_preflight_files=preview["planned_files"],
                apply=True,
            )

    assert not tuple(public.iterdir())


def test_apply_rejects_existing_symlinked_selected_media(api, tmp_path: Path) -> None:
    repair = _repair_module()
    rows = [("article-1", "12621604", ".zip")]
    articles, source, public, private = _prepare_source(tmp_path, rows)
    outside = tmp_path / "outside.zip"
    outside.write_bytes(_zip_bytes())
    selected_link = public / "legacy-12621604.zip"
    try:
        selected_link.symlink_to(outside)
    except OSError as exc:
        pytest.skip(f"symlink creation is unavailable: {exc}")

    with api.session() as db:
        post = db.get(Post, 3)
        post.status = "published"
        post.deleted_at = None
        post.metadata_json = {"legacy_write_id": "article-1"}
        db.commit()

        with pytest.raises(ValueError, match="symlinked selected media file is not allowed"):
            repair.repair_legacy_attachments(
                db,
                articles_xlsx=articles,
                attachment_source_dir=source,
                public_media_dir=public,
                private_media_dir=private,
                storage_ids=("12621604",),
                apply=True,
            )

    assert selected_link.is_symlink()
    assert outside.read_bytes() == _zip_bytes()


def test_rollback_expands_user_media_paths(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    repair = _repair_module()
    monkeypatch.setenv("HOME", str(tmp_path))
    monkeypatch.setenv("USERPROFILE", str(tmp_path))
    public = tmp_path / "public"
    private = tmp_path / "private"
    public.mkdir()
    private.mkdir()
    selected = public / "legacy-12621604.zip"
    selected.write_bytes(b"original")

    with pytest.raises(RuntimeError, match="injected rollback"):
        with repair._restore_media_files_on_error(
            Path("~/public"),
            Path("~/private"),
            ("12621604",),
        ):
            selected.write_bytes(b"changed")
            raise RuntimeError("injected rollback")

    assert selected.read_bytes() == b"original"


def test_apply_rejects_existing_media_that_does_not_match_preflight_source(api, tmp_path: Path) -> None:
    repair = _repair_module()
    rows = [("article-1", "10946091", ".doc")]
    articles, source, public, private = _prepare_source(tmp_path, rows)
    word_body = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1" + (b"\x00" * 256)
    old_hwp = public / "legacy-10946091.doc"
    old_hwp.write_bytes(word_body)

    with api.session() as db:
        post = db.get(Post, 3)
        post.metadata_json = {"legacy_write_id": "article-1"}
        media = MediaAsset(
            owner_id=post.author_id,
            original_filename="legacy-10946091.doc",
            stored_filename=old_hwp.name,
            content_type="application/msword",
            file_size=len(word_body),
            url=None,
            is_private=False,
            status="ready",
        )
        db.add(media)
        db.flush()
        media.url = f"/api/media/{media.id}/access-url"
        db.add(PostAttachment(post_id=post.id, media_id=media.id, sort_order=0))
        db.commit()

        with pytest.raises(RuntimeError, match="postcondition"):
            repair.repair_legacy_attachments(
                db,
                articles_xlsx=articles,
                attachment_source_dir=source,
                public_media_dir=public,
                private_media_dir=private,
                storage_ids=("10946091",),
                apply=True,
            )
        db.refresh(media)
        assert media.stored_filename == "legacy-10946091.doc"
        assert media.content_type == "application/msword"

    assert old_hwp.read_bytes() == word_body
    assert not (public / "legacy-10946091.hwp").exists()


def test_failed_selected_repair_rolls_back_database_and_media_files(
    api,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    repair = _repair_module()
    rows = [
        ("article-1", "10946091", ".hwp"),
        ("article-1", "12621604", ".zip"),
    ]
    articles, source, public, private = _prepare_source(tmp_path, rows)
    old_hwp = public / "legacy-10946091.doc"
    old_hwp.write_bytes(_source_bytes(".hwp"))

    with api.session() as db:
        post = db.get(Post, 3)
        post.metadata_json = {"legacy_write_id": "article-1"}
        media = MediaAsset(
            owner_id=post.author_id,
            original_filename="legacy-10946091",
            stored_filename=old_hwp.name,
            content_type="application/msword",
            file_size=old_hwp.stat().st_size,
            url=None,
            is_private=False,
            status="ready",
        )
        db.add(media)
        db.flush()
        media.url = f"/api/media/{media.id}/access-url"
        db.add(PostAttachment(post_id=post.id, media_id=media.id, sort_order=0))
        db.commit()

        original_insert = repair._insert_missing_attachments

        def fail_after_insert(*args, **kwargs):
            original_insert(*args, **kwargs)
            raise RuntimeError("injected post-import failure")

        monkeypatch.setattr(repair, "_insert_missing_attachments", fail_after_insert)

        with pytest.raises(RuntimeError, match="injected post-import failure"):
            repair.repair_legacy_attachments(
                db,
                articles_xlsx=articles,
                attachment_source_dir=source,
                public_media_dir=public,
                private_media_dir=private,
                storage_ids=("10946091", "12621604"),
                apply=True,
            )
        db.refresh(media)

        assert media.stored_filename == "legacy-10946091.doc"
        assert media.content_type == "application/msword"

    assert old_hwp.read_bytes() == _source_bytes(".hwp")
    assert not (public / "legacy-10946091.hwp").exists()
    assert not (public / "legacy-12621604.zip").exists()
