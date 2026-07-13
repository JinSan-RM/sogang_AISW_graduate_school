from __future__ import annotations

import argparse
import csv
import hashlib
import math
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime, timezone
from html import unescape
from html.parser import HTMLParser
from mimetypes import guess_type
from pathlib import Path
from typing import Any

from sqlalchemy import select
from sqlalchemy.orm import Session

BACKEND_ROOT = Path(__file__).resolve().parents[1]
REPO_ROOT = BACKEND_ROOT.parent
if str(BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(BACKEND_ROOT))

from app.database import SessionLocal  # noqa: E402
from app.models.board import Board  # noqa: E402
from app.models.comment import Comment  # noqa: E402
from app.models.media import MediaAsset, PostAttachment  # noqa: E402
from app.models.post import Post  # noqa: E402
from app.models.post_extension import PostSuggestion  # noqa: E402
from app.models.user import User  # noqa: E402
from app.security import hash_password  # noqa: E402
from app.seed import ACTIVE_BOARD_SEED_DATA, seed_initial_data  # noqa: E402

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - surfaced as a CLI blocker
    raise SystemExit("openpyxl is required. Install backend requirements before running this importer.") from exc


SHEET_TO_BOARD_SLUG = {
    "전체공지": "all-notices",
    "학사일정": "academic-calendar",
    "전공 커뮤니티": "community-major",
    "강의후기": "lecture-reviews",
    "자료 공유": "exam-archive",
    "건의 사항": "suggestions",
    "건의사항 피드백": "gsa-feedback",
    "자주묻는질문": "gsa-faq",
    "회계 장부": "accounting",
    "임원진 소개": "gsa-executives",
    "사진첩": "event-album",
    "세미나 공유": "community-seminar",
    "스터디 모집": "study-recruit",
    "스터디 지원 신청": "study-apply",
    "원우회 상조회": "mutual-aid",
    "동아리 지원 신청": "club-apply",
    "동아리 홍보": "club-promo",
    "원우회 활동내역": "council-activity",
    "기장단": "gsa-cohort-leaders",
    "종합시험": "comprehensive-exam",
    "로드맵_원우회비혜택": "gsa-roadmap-benefits",
    "웨비나_특강": "webinar-notices",
    "채용정보": "community-job",
    "논문 자료": "community-paper",
    "동문 주소록": "alumni-directory",
}

CSV_LABEL_TO_BOARD_SLUG = {
    "notice": "all-notices",
    "lecture_review": "lecture-reviews",
    "shared_resources": "exam-archive",
    "major_community": "community-major",
    "suggestions": "suggestions",
    "faq": "gsa-faq",
}

BOARD_LABEL_BY_SLUG = {item["slug"]: item["name"] for item in ACTIVE_BOARD_SEED_DATA}

ATTACHMENT_SHEET = "첨부파일"
COMMENT_RE = re.compile(
    r"^\[(?P<created_at>[^\]]+)\]\s*(?P<author>[^:\n]+):\s*(?P<body>.*?)(?=^\[[0-9]{4}-|\Z)",
    re.MULTILINE | re.DOTALL,
)


class HTMLTextExtractor(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.parts: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag in {"br", "p", "div", "li", "tr"}:
            self.parts.append("\n")
        if tag == "a":
            href = dict(attrs).get("href")
            if href:
                self.parts.append(f" {href} ")

    def handle_endtag(self, tag: str) -> None:
        if tag in {"p", "div", "li", "tr"}:
            self.parts.append("\n")

    def handle_data(self, data: str) -> None:
        self.parts.append(data)

    def text(self) -> str:
        return "".join(self.parts)


def resolve_default(pattern: str) -> Path | None:
    matches = sorted(REPO_ROOT.glob(pattern))
    return matches[0] if matches else None


def cell_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    return value


def value_as_text(value: Any) -> str:
    value = cell_value(value)
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def value_as_int(value: Any, default: int = 0) -> int:
    value = cell_value(value)
    if value is None or value == "":
        return default
    try:
        return int(float(str(value).replace(",", "")))
    except ValueError:
        return default


def value_as_bool(value: Any) -> bool:
    value = cell_value(value)
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    return str(value).strip().lower() in {"1", "true", "yes", "y"}


def parse_datetime(value: Any) -> datetime:
    value = cell_value(value)
    if isinstance(value, datetime):
        parsed = value
    elif value:
        text = str(value).strip().replace("Z", "+00:00")
        parsed = None
        for date_format in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y.%m.%d %H:%M:%S", "%Y.%m.%d %H:%M"):
            try:
                parsed = datetime.strptime(text, date_format)
                break
            except ValueError:
                pass
        if parsed is None:
            parsed = datetime.fromisoformat(text)
    else:
        parsed = datetime.utcnow()
    if parsed.tzinfo is not None:
        parsed = parsed.astimezone(timezone.utc).replace(tzinfo=None)
    return parsed


def normalize_content(value: Any) -> tuple[str, bool]:
    text = value_as_text(value)
    looks_like_html = "<" in text and ">" in text
    if not looks_like_html:
        return text, False

    parser = HTMLTextExtractor()
    parser.feed(text)
    extracted = unescape(parser.text())
    extracted = re.sub(r"[ \t]+\n", "\n", extracted)
    extracted = re.sub(r"\n{3,}", "\n\n", extracted)
    return extracted.strip(), True


def legacy_id(value: Any) -> str:
    text = value_as_text(value)
    if text.endswith(".0"):
        text = text[:-2]
    return text


def legacy_user_key(nickname: str) -> tuple[str, str]:
    digest = hashlib.sha1(nickname.encode("utf-8")).hexdigest()[:20]
    return f"legacy_{digest}", f"legacy-{digest}@legacy.local"


def infer_cohort(nickname: str) -> str | None:
    match = re.search(r"(\d{2})\s*기|^(\d{2})[_-]", nickname)
    if not match:
        return None
    return next(group for group in match.groups() if group)


def get_or_create_legacy_user(db: Session, nickname: str, cache: dict[str, User]) -> User:
    nickname = nickname.strip() or "레거시 작성자"
    if nickname in cache:
        return cache[nickname]

    username, email = legacy_user_key(nickname)
    user = db.scalar(select(User).where(User.username == username))
    if user is None:
        user = User(
            username=username,
            email=email,
            password_hash=hash_password(f"legacy:{username}"),
            nickname=nickname[:50],
            cohort=infer_cohort(nickname),
            role="user",
            is_active=False,
        )
        db.add(user)
        db.flush()
    cache[nickname] = user
    return user


def iter_sheet_records(workbook_path: Path) -> tuple[dict[str, list[dict[str, Any]]], dict[str, list[dict[str, Any]]]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    article_records: dict[str, list[dict[str, Any]]] = {}
    attachments_by_write_id: dict[str, list[dict[str, Any]]] = defaultdict(list)

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        rows = sheet.iter_rows(values_only=True)
        headers = [value_as_text(value) for value in next(rows, [])]
        if not headers:
            continue

        for values in rows:
            record = {headers[index]: cell_value(value) for index, value in enumerate(values) if index < len(headers)}
            if all(value is None or value == "" for value in record.values()):
                continue
            if sheet_name == ATTACHMENT_SHEET:
                write_id = legacy_id(record.get("writeId"))
                if write_id:
                    attachments_by_write_id[write_id].append(record)
            else:
                article_records.setdefault(sheet_name, []).append(record)

    return article_records, attachments_by_write_id


def existing_posts_by_legacy_id(db: Session) -> dict[str, Post]:
    posts = db.scalars(select(Post)).all()
    result: dict[str, Post] = {}
    for post in posts:
        metadata = post.metadata_json or {}
        write_id = metadata.get("legacy_write_id")
        if write_id is not None:
            result[str(write_id)] = post
    return result


def existing_comment_keys(db: Session) -> set[tuple[int, int, datetime, str]]:
    comments = db.scalars(select(Comment)).all()
    return {(comment.post_id, comment.author_id, comment.created_at, comment.content[:120]) for comment in comments}


def attachment_urls_for_post(db: Session, post_id: int) -> set[str]:
    rows = db.execute(
        select(MediaAsset.url)
        .join(PostAttachment, PostAttachment.media_id == MediaAsset.id)
        .where(PostAttachment.post_id == post_id)
    ).all()
    return {url for (url,) in rows if url}


def add_attachments(
    db: Session,
    post: Post,
    attachment_records: list[dict[str, Any]],
    existing_urls: set[str],
) -> int:
    added = 0
    for index, attachment in enumerate(attachment_records):
        url = value_as_text(attachment.get("url"))
        if not url or url in existing_urls:
            continue
        filename = value_as_text(attachment.get("subject")) or Path(url).name or f"legacy-{post.id}-{index}"
        storage_id = value_as_text(attachment.get("fileStorageId")) or Path(url).name or filename
        guessed_type, _ = guess_type(filename)
        media = MediaAsset(
            owner_id=post.author_id,
            original_filename=filename[:255],
            stored_filename=storage_id[:255],
            content_type=(guessed_type or value_as_text(attachment.get("contentType")) or "application/octet-stream")[:100],
            file_size=value_as_int(attachment.get("fileSize")),
            url=url[:500],
            status="ready",
            created_at=parse_datetime(attachment.get("regiDatetime")),
        )
        db.add(media)
        db.flush()
        db.add(PostAttachment(post_id=post.id, media_id=media.id, sort_order=index))
        existing_urls.add(url)
        added += 1
    return added


def import_articles(
    db: Session,
    workbook_path: Path,
    *,
    limit: int | None,
    update_existing: bool,
) -> dict[str, int]:
    article_records, attachments_by_write_id = iter_sheet_records(workbook_path)
    board_by_slug = {board.slug: board for board in db.scalars(select(Board)).all()}
    existing_posts = existing_posts_by_legacy_id(db)
    user_cache: dict[str, User] = {}
    stats = Counter()

    for sheet_name, records in article_records.items():
        slug = SHEET_TO_BOARD_SLUG.get(sheet_name)
        if not slug:
            stats["skipped_unmapped_sheet_rows"] += len(records)
            continue
        board = board_by_slug.get(slug)
        if board is None:
            stats["skipped_missing_board_rows"] += len(records)
            continue

        for record in records:
            if limit is not None and stats["processed_rows"] >= limit:
                return dict(stats)
            write_id = legacy_id(record.get("writeId"))
            if not write_id:
                stats["skipped_missing_write_id"] += 1
                continue

            title = value_as_text(record.get("title")) or "(제목 없음)"
            content, content_was_html = normalize_content(record.get("content"))
            if not content:
                content = title
            author = get_or_create_legacy_user(db, value_as_text(record.get("writeUser")), user_cache)
            attachments = attachments_by_write_id.get(write_id, [])
            metadata = {
                "legacy_source": "swing2app",
                "legacy_sheet": sheet_name,
                "legacy_write_id": write_id,
                "legacy_board_id": value_as_text(record.get("boardId")),
                "legacy_parent_write_id": value_as_text(record.get("parentWriteId")) or None,
                "legacy_write_type": value_as_text(record.get("writeType")),
                "legacy_author": value_as_text(record.get("writeUser")),
                "legacy_stat": value_as_text(record.get("stat")),
                "legacy_notices_missing": value_as_bool(record.get("notices_미포함")),
                "legacy_comment_count": value_as_int(record.get("commentCount")),
                "legacy_attachment_count": len(attachments),
                "legacy_content_was_html": content_was_html,
            }

            post = existing_posts.get(write_id)
            if post is None:
                post = Post(
                    board_id=board.id,
                    author_id=author.id,
                    title=title[:200],
                    content=content,
                    is_anonymous=False,
                    is_notice=board.board_type == "notice",
                    status="published",
                    category=sheet_name,
                    metadata_json=metadata,
                    view_count=value_as_int(record.get("viewCount")),
                    like_count=value_as_int(record.get("likeCount")),
                    comment_count=value_as_int(record.get("commentCount")),
                    created_at=parse_datetime(record.get("date")),
                    updated_at=parse_datetime(record.get("updateDate")),
                )
                db.add(post)
                db.flush()
                existing_posts[write_id] = post
                stats["created_posts"] += 1
            elif update_existing:
                post.board_id = board.id
                post.author_id = author.id
                post.title = title[:200]
                post.content = content
                post.is_notice = board.board_type == "notice"
                post.category = sheet_name
                post.metadata_json = {**(post.metadata_json or {}), **metadata}
                post.view_count = value_as_int(record.get("viewCount"))
                post.like_count = value_as_int(record.get("likeCount"))
                post.comment_count = max(post.comment_count, value_as_int(record.get("commentCount")))
                post.created_at = parse_datetime(record.get("date"))
                post.updated_at = parse_datetime(record.get("updateDate"))
                stats["updated_posts"] += 1
            else:
                stats["skipped_existing_posts"] += 1

            if board.board_type == "suggestion":
                suggestion = db.scalar(select(PostSuggestion).where(PostSuggestion.post_id == post.id))
                if suggestion is None:
                    db.add(PostSuggestion(post_id=post.id, suggestion_category=sheet_name, status="received"))

            stats["created_attachments"] += add_attachments(db, post, attachments, attachment_urls_for_post(db, post.id))
            stats["processed_rows"] += 1

    return dict(stats)


def parse_comments(raw_comments: str) -> list[dict[str, Any]]:
    comments = []
    for match in COMMENT_RE.finditer(raw_comments):
        body = match.group("body").strip()
        if not body:
            continue
        comments.append(
            {
                "created_at": parse_datetime(match.group("created_at")),
                "author": match.group("author").strip(),
                "content": body,
            }
        )
    return comments


def import_comments(db: Session, comments_csv_path: Path) -> dict[str, int]:
    posts_by_legacy_id = existing_posts_by_legacy_id(db)
    existing_keys = existing_comment_keys(db)
    user_cache: dict[str, User] = {}
    stats = Counter()

    with comments_csv_path.open("r", encoding="utf-8-sig", newline="") as file:
        reader = csv.DictReader(file)
        for row in reader:
            raw_comments = value_as_text(row.get("comments"))
            if not raw_comments:
                continue
            post = posts_by_legacy_id.get(legacy_id(row.get("post_id")))
            if post is None:
                stats["skipped_missing_post"] += 1
                continue
            parsed_comments = parse_comments(raw_comments)
            for comment_payload in parsed_comments:
                author = get_or_create_legacy_user(db, comment_payload["author"], user_cache)
                key = (post.id, author.id, comment_payload["created_at"], comment_payload["content"][:120])
                if key in existing_keys:
                    stats["skipped_existing_comments"] += 1
                    continue
                db.add(
                    Comment(
                        post_id=post.id,
                        author_id=author.id,
                        parent_id=None,
                        content=comment_payload["content"],
                        created_at=comment_payload["created_at"],
                        updated_at=comment_payload["created_at"],
                    )
                )
                existing_keys.add(key)
                stats["created_comments"] += 1
            post.comment_count = max(post.comment_count, len(parsed_comments))
    return dict(stats)


def dry_run(workbook_path: Path, comments_csv_path: Path | None) -> None:
    article_records, attachments_by_write_id = iter_sheet_records(workbook_path)
    print(f"Workbook: {workbook_path}")
    print(f"Article sheets: {len(article_records)}")
    total_rows = 0
    for sheet_name, records in article_records.items():
        total_rows += len(records)
        slug = SHEET_TO_BOARD_SLUG.get(sheet_name)
        target = f"{BOARD_LABEL_BY_SLUG.get(slug, slug)} ({slug})" if slug else "UNMAPPED"
        print(f"- {sheet_name}: {len(records)} rows -> {target}")
    print(f"Article rows: {total_rows}")
    print(f"Attachment rows: {sum(len(items) for items in attachments_by_write_id.values())}")
    print(f"Posts with attachments: {len(attachments_by_write_id)}")

    if comments_csv_path and comments_csv_path.exists():
        comment_rows = 0
        parsed_comments = 0
        with comments_csv_path.open("r", encoding="utf-8-sig", newline="") as file:
            reader = csv.DictReader(file)
            for row in reader:
                comments = parse_comments(value_as_text(row.get("comments")))
                if comments:
                    comment_rows += 1
                    parsed_comments += len(comments)
        print(f"CSV rows with comments: {comment_rows}")
        print(f"Parsed comments: {parsed_comments}")


def main() -> None:
    default_workbook = resolve_default("board_articles*.xlsx")
    default_csv = resolve_default("*작성글*.csv")
    parser = argparse.ArgumentParser(description="Import legacy Swing2App articles into the local app database.")
    parser.add_argument("--workbook", type=Path, default=default_workbook, help="Path to board_articles workbook.")
    parser.add_argument("--comments-csv", type=Path, default=default_csv, help="Optional CSV with legacy comments.")
    parser.add_argument("--dry-run", action="store_true", help="Parse source files and print an import summary without DB writes.")
    parser.add_argument("--limit", type=int, default=None, help="Limit article rows for a test import.")
    parser.add_argument("--no-update-existing", action="store_true", help="Skip posts already imported by legacy writeId.")
    parser.add_argument("--skip-comments", action="store_true", help="Do not import comments from the CSV file.")
    args = parser.parse_args()

    if args.workbook is None or not args.workbook.exists():
        raise SystemExit("Could not find board_articles workbook. Pass --workbook <path>.")

    if args.dry_run:
        dry_run(args.workbook, args.comments_csv if args.comments_csv and args.comments_csv.exists() else None)
        return

    db = SessionLocal()
    try:
        seed_initial_data(db)
        article_stats = import_articles(
            db,
            args.workbook,
            limit=args.limit,
            update_existing=not args.no_update_existing,
        )
        comment_stats = {}
        if not args.skip_comments and args.comments_csv and args.comments_csv.exists():
            comment_stats = import_comments(db, args.comments_csv)
        db.commit()
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()

    print("Article import:", article_stats)
    if comment_stats:
        print("Comment import:", comment_stats)


if __name__ == "__main__":
    main()
